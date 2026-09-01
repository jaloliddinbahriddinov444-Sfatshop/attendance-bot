"""Payroll (ish haqqi Excel) hisobotiga kim kirishi testi.

Qoida: faqat role='employee'. Rahbariyat (boss / bosh_admin) kunbay
hisoblanmaydi, shuning uchun hisobotda umuman ko'rinmaydi.

Ishga tushirish: ./venv/bin/python test_payroll_roles.py
"""
import os
import tempfile

os.environ["DB_PATH"] = os.path.join(tempfile.mkdtemp(), "payroll_test.db")
os.environ.setdefault("BOT_TOKEN", "test:token")

import database as db          # noqa: E402

Y, M = 2026, 9


def _mk(tg: int, name: str, role: str, rate: int) -> int:
    with db.get_db() as conn:
        cur = conn.execute(
            """INSERT INTO employees (telegram_id, full_name, phone, position,
               face_encoding, role, is_admin, is_active, position_id, daily_rate,
               registered_at)
               VALUES (?,?,?,'TestLavozim',x'00',?,0,1,1,?,'2020-01-01 00:00:00')""",
            (tg, name, f"+9989{tg}", role, rate)
        )
        return cur.lastrowid


def main():
    db.init_db()
    with db.get_db() as conn:
        conn.execute("INSERT OR IGNORE INTO positions (id, name, work_hours) VALUES (1,'Sotuvchi',9)")

    _mk(1001, "Oddiy Xodim", "employee", 150000)
    _mk(1002, "Bosh Admin", "bosh_admin", 150000)
    _mk(1003, "Rahbar", "boss", 150000)
    _mk(1004, "Demo Rahbar", "boss", 150000)

    # Bayram — rahbariyatda ham summa paydo bo'lishi mumkin, baribir kirmasin
    db.set_calendar_day(f"{Y}-{M:02d}-10", db.HOLIDAY, 1)

    names = [r["employee"]["full_name"]
             for r in db.get_all_employees_salary_summary(Y, M)]
    assert names == ["Oddiy Xodim"], names
    print("✅ payroll hisobotida faqat role='employee'")

    # Bayram haqqi oddiy xodimga baribir yoziladi
    emp = db.get_employee_by_telegram_id(1001)
    assert db.get_monthly_holiday_pay(emp["id"], Y, M) == 150000
    assert db.get_holiday_pay_days(emp["id"], Y, M) == [f"{Y}-{M:02d}-10"]
    row = db.get_all_employees_salary_summary(Y, M)[0]
    assert row["base"] == 150000, row["base"]
    assert row["holiday"] == 150000, row["holiday"]
    print("✅ bayram haqqi hisobot qatoriga tushdi")

    # role='admin' — haqiqiy xodim, payrolldan tushib qolmasin
    _mk(1005, "Admin Xodim", "admin", 120000)
    names = [r["employee"]["full_name"]
             for r in db.get_all_employees_salary_summary(Y, M)]
    assert "Admin Xodim" in names, names
    print("✅ 'admin' roli payrollda qoladi")

    # Excelda ko'rinadigan bo'ldimi: kunlar jadvali + «shundan bayram» qatori
    import sys, types, io
    fake = types.ModuleType("face_recognition")
    fake.face_encodings = lambda *a, **k: []
    sys.modules.setdefault("face_recognition", fake)
    from handlers.emp_data import _build_emp_excel
    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(_build_emp_excel(emp["id"], Y, M))).active
    cells = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    flat = [str(c) for row in cells for c in row if c is not None]
    assert any("🎉 Bayram kuni" in s for s in flat), "kunlar jadvalida bayram yo'q"
    assert any("Shundan bayram" in s for s in flat), "xulosada bayram qatori yo'q"
    print("✅ xodim Excelida bayram kuni va «shundan bayram» ko'rinadi")

    print("\n✅✅ HAMMASI O'TDI — jonli attendance.db tegilmadi")


if __name__ == "__main__":
    main()
