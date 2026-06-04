"""Admin paneli: xodim boshqaruvi, sozlamalar, hisobotlar"""
import asyncio
import io
import logging
from datetime import datetime
from tzutil import now as tz_now, to_local
from aiogram import Router, F, Bot
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, BufferedInputFile, InlineKeyboardMarkup, InlineKeyboardButton

import texts
import keyboards as kb
from states import AdminPanel, AdminSalary, TaskCreate, AdminAddEmployee
from database import (
    get_employee_by_telegram_id, get_all_employees, get_employee_by_id,
    deactivate_employee, set_admin_status, get_active_employees_count,
    get_today_all_attendance, get_office_config, set_setting,
    get_monthly_attendance, delete_today_attendance, add_manual_attendance,
    set_hourly_rate, add_salary_entry, cancel_salary_entry,
    get_salary_entry, get_active_salary_entries,
    is_month_closed, close_month, reopen_month,
    get_audit_entries, get_all_employees_salary_summary,
    get_monthly_worked_minutes, get_salary_totals_by_type,
    create_task, set_role, get_boss,
    find_employee_by_phone, create_pending_employee,
    update_employee_profile, reactivate_employee, phone_key,
)
from config import MAX_EMPLOYEES

logger = logging.getLogger(__name__)
router = Router()


def _admin_kb(actor) -> object:
    """Hozirgi foydalanuvchiga mos keladigan menyu (Bosh Admin/Admin/Boss)."""
    uid = actor.from_user.id
    emp = get_employee_by_telegram_id(uid)
    role = "employee"
    if emp:
        try:
            role = emp["role"] or "employee"
        except (KeyError, IndexError):
            role = "admin" if emp["is_admin"] else "employee"
    if role == "boss":
        return kb.boss_panel_kb()
    return kb.admin_menu_kb(is_bosh_admin=(role == "bosh_admin"))


def _is_admin(message: Message) -> bool:
    emp = get_employee_by_telegram_id(message.from_user.id)
    return emp is not None and bool(emp["is_admin"])


@router.message(F.text == texts.BTN_ADMIN)
async def admin_menu(message: Message, state: FSMContext):
    if not _is_admin(message):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()
    await message.answer(texts.ADMIN_MENU, reply_markup=_admin_kb(message))


# ===== Bosh Admin: bo'lim submenyulari =====

def _is_bosh_admin(message: Message) -> bool:
    emp = get_employee_by_telegram_id(message.from_user.id)
    if not emp:
        return False
    try:
        return (emp["role"] or "") == "bosh_admin"
    except (KeyError, IndexError):
        return False


@router.message(F.text == texts.BTN_GRP_EMPLOYEES)
async def grp_employees(message: Message, state: FSMContext):
    if not _is_bosh_admin(message):
        return
    await state.clear()
    await message.answer(texts.GRP_EMPLOYEES_HEADER, reply_markup=kb.grp_employees_kb())


@router.message(F.text == texts.BTN_GRP_ATTENDANCE)
async def grp_attendance(message: Message, state: FSMContext):
    if not _is_bosh_admin(message):
        return
    await state.clear()
    await message.answer(texts.GRP_ATTENDANCE_HEADER, reply_markup=kb.grp_attendance_kb())


@router.message(F.text == texts.BTN_GRP_FINANCE)
async def grp_finance(message: Message, state: FSMContext):
    if not _is_bosh_admin(message):
        return
    await state.clear()
    await message.answer(texts.GRP_FINANCE_HEADER, reply_markup=kb.grp_finance_kb())


@router.message(F.text == texts.BTN_GRP_CONTROL)
async def grp_control(message: Message, state: FSMContext):
    if not _is_bosh_admin(message):
        return
    await state.clear()
    await message.answer(texts.GRP_CONTROL_HEADER, reply_markup=kb.grp_control_kb())


@router.message(F.text == texts.BTN_ADMIN_BACK)
async def grp_back_to_panel(message: Message, state: FSMContext):
    if not _is_admin(message):
        return
    await state.clear()
    await message.answer(texts.ADMIN_MENU, reply_markup=_admin_kb(message))


# ===== Phase 4: Admin xodim qo'shish (oldindan) =====

@router.message(F.text == texts.BTN_ADMIN_ADD_EMPLOYEE)
async def add_emp_start(message: Message, state: FSMContext):
    if not _is_admin(message):
        return
    await state.clear()
    await message.answer(texts.ADD_EMP_ASK_NAME, reply_markup=kb.cancel_kb())
    await state.set_state(AdminAddEmployee.waiting_name)


@router.message(AdminAddEmployee.waiting_name, F.text == texts.BTN_CANCEL)
@router.message(AdminAddEmployee.waiting_phone, F.text == texts.BTN_CANCEL)
@router.message(AdminAddEmployee.waiting_position, F.text == texts.BTN_CANCEL)
async def add_emp_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(texts.CANCELLED, reply_markup=_admin_kb(message))


@router.message(AdminAddEmployee.waiting_name, F.text)
async def add_emp_name(message: Message, state: FSMContext):
    name = message.text.strip()
    if len(name) < 5:
        await message.answer(texts.ADD_EMP_NAME_TOO_SHORT)
        return
    await state.update_data(ae_name=name)
    await message.answer(texts.ADD_EMP_ASK_PHONE, reply_markup=kb.cancel_kb())
    await state.set_state(AdminAddEmployee.waiting_phone)


@router.message(AdminAddEmployee.waiting_phone, F.text)
async def add_emp_phone(message: Message, state: FSMContext):
    phone = message.text.strip()
    if len(phone_key(phone)) < 9:
        await message.answer(texts.ADD_EMP_PHONE_INVALID)
        return
    await state.update_data(ae_phone=phone)
    await message.answer(texts.ADD_EMP_ASK_POSITION, reply_markup=kb.cancel_kb())
    await state.set_state(AdminAddEmployee.waiting_position)


@router.message(AdminAddEmployee.waiting_position, F.text)
async def add_emp_position(message: Message, state: FSMContext):
    position = message.text.strip()
    if len(position) < 2:
        await message.answer(texts.ADD_EMP_POSITION_TOO_SHORT)
        return
    data = await state.update_data(ae_position=position)
    await message.answer(
        texts.ADD_EMP_CONFIRM.format(
            name=data["ae_name"], phone=data["ae_phone"], position=position
        ),
        reply_markup=kb.addemp_confirm_kb()
    )
    await state.set_state(AdminAddEmployee.waiting_confirm)


@router.callback_query(F.data.startswith("addemp:"))
async def add_emp_confirm(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    answer = call.data.split(":")[1]
    if answer == "no":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    data = await state.get_data()
    name = data.get("ae_name")
    phone = data.get("ae_phone")
    position = data.get("ae_position")
    await state.clear()

    if not (name and phone and position):
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    existing = find_employee_by_phone(phone, include_inactive=True)

    if existing and existing["telegram_id"] > 0 and existing["is_active"]:
        # Faol va bog'langan — qo'shib bo'lmaydi
        await call.message.edit_text(
            texts.ADD_EMP_ALREADY_ACTIVE.format(name=existing["full_name"])
        )
        await call.answer()
        return

    if existing:
        # Pending (telegram_id<0) yoki deaktiv yozuv — jonlantirib yangilaymiz
        if not existing["is_active"]:
            reactivate_employee(existing["id"])
        update_employee_profile(existing["id"], full_name=name, position=position)
        if existing["telegram_id"] < 0:
            msg = texts.ADD_EMP_UPDATED.format(name=name, phone=phone)
        else:
            msg = texts.ADD_EMP_REACTIVATED.format(name=name, phone=phone)
        await call.message.edit_text(msg)
        await call.answer("✅")
        logger.info("Add-emp: existing id=%s updated by tg=%s",
                    existing["id"], call.from_user.id)
        return

    # Yangi pending yozuv — limit tekshiruvi
    if get_active_employees_count() >= MAX_EMPLOYEES:
        await call.message.edit_text(texts.ADD_EMP_LIMIT.format(max=MAX_EMPLOYEES))
        await call.answer()
        return

    new_id = create_pending_employee(name, phone, position)
    await call.message.edit_text(texts.ADD_EMP_ADDED.format(name=name, phone=phone))
    await call.answer("✅")
    logger.info("Add-emp: pending created id=%s by tg=%s", new_id, call.from_user.id)


# ===== Xodimlar ro'yxati =====

@router.message(F.text == texts.BTN_ADMIN_LIST)
async def admin_list(message: Message, bot: Bot):
    if not _is_admin(message):
        return
    employees = get_all_employees(active_only=True)
    count = len(employees)

    me = await bot.get_me()
    text = texts.ADMIN_INVITE_LINK.format(
        bot_username=me.username, count=count, max=MAX_EMPLOYEES
    )
    text += "\n\n" + texts.EMPLOYEES_LIST_HEADER.format(count=count, max=MAX_EMPLOYEES)

    if not employees:
        text += "<i>Xodimlar yo'q</i>"
    else:
        for idx, emp in enumerate(employees, 1):
            text += texts.EMPLOYEE_ITEM.format(
                idx=idx,
                admin_icon="👑" if emp["is_admin"] else "👤",
                name=emp["full_name"],
                position=emp["position"],
                phone=emp["phone"]
            )

    await message.answer(text)


# ===== Xodim o'chirish =====

@router.message(F.text == texts.BTN_ADMIN_REMOVE)
async def admin_remove_start(message: Message, state: FSMContext):
    if not _is_admin(message):
        return
    employees = get_all_employees(active_only=True)
    if not employees:
        await message.answer("📋 Xodimlar ro'yxati bo'sh.")
        return
    await message.answer(
        texts.ADMIN_REMOVE_PROMPT,
        reply_markup=kb.employees_inline_kb(employees, "remove")
    )


@router.callback_query(F.data.startswith("remove:"))
async def admin_remove_callback(call: CallbackQuery, bot: Bot):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    parts = call.data.split(":")
    action = parts[1]

    if action == "cancel":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    # remove:<id>
    if len(parts) == 2:
        emp_id = int(parts[1])
        emp = get_employee_by_id(emp_id)
        if not emp:
            await call.answer("❌ Topilmadi", show_alert=True)
            return
        await call.message.edit_text(
            texts.ADMIN_REMOVE_CONFIRM.format(name=emp["full_name"]),
            reply_markup=kb.confirm_inline_kb("remove_confirm", emp_id)
        )
        await call.answer()


@router.callback_query(F.data.startswith("remove_confirm:"))
async def admin_remove_confirm(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    _, answer, emp_id = call.data.split(":")
    emp_id = int(emp_id)
    emp = get_employee_by_id(emp_id)

    if answer == "no" or not emp:
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    # O'zini o'chirishga ruxsat bermaymiz
    if emp["telegram_id"] == call.from_user.id:
        await call.answer("❌ O'zingizni o'chira olmaysiz", show_alert=True)
        return

    deactivate_employee(emp_id)
    await call.message.edit_text(texts.ADMIN_REMOVE_DONE.format(name=emp["full_name"]))
    await call.answer("✅ O'chirildi")


# ===== Admin tayinlash =====

@router.message(F.text == texts.BTN_ADMIN_PROMOTE)
async def admin_promote_start(message: Message):
    if not _is_admin(message):
        return
    employees = get_all_employees(active_only=True)
    if not employees:
        await message.answer("📋 Xodimlar ro'yxati bo'sh.")
        return
    await message.answer(
        texts.ADMIN_PROMOTE_PROMPT,
        reply_markup=kb.employees_inline_kb(employees, "promote")
    )


@router.callback_query(F.data.startswith("promote:"))
async def admin_promote_callback(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    parts = call.data.split(":")
    if parts[1] == "cancel":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    emp_id = int(parts[1])
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    new_status = not bool(emp["is_admin"])
    set_admin_status(emp_id, new_status)
    status_text = "👑 admin qilindi" if new_status else "👤 oddiy xodim qilindi"
    await call.message.edit_text(
        f"✅ <b>{emp['full_name']}</b> {status_text}."
    )
    await call.answer()


# ===== Bugungi hisobot =====

@router.message(F.text == texts.BTN_ADMIN_TODAY)
async def admin_today(message: Message):
    if not _is_admin(message):
        return

    records = get_today_all_attendance()
    today_str = tz_now().strftime("%d.%m.%Y")
    text = texts.ADMIN_TODAY_HEADER.format(date=today_str)

    if not records:
        text += "<i>Xodimlar yo'q</i>"
    else:
        for rec in records:
            if rec["first_in"]:
                out_line = (texts.OUT_LINE_PRESENT.format(out_time=rec["last_out"][:5])
                            if rec["last_out"] else texts.OUT_LINE_ABSENT)
                wifi_warn = texts.WIFI_WARN if rec["has_wifi_warning"] else ""
                text += texts.ADMIN_TODAY_ITEM_PRESENT.format(
                    name=rec["full_name"],
                    in_time=rec["first_in"][:5],
                    out_line=out_line,
                    wifi_warn=("   " + wifi_warn) if wifi_warn else ""
                )
            else:
                text += texts.ADMIN_TODAY_ITEM_ABSENT.format(name=rec["full_name"])

    await message.answer(text)


# ===== Excel hisobot =====

@router.message(F.text == texts.BTN_ADMIN_EXPORT)
async def admin_export(message: Message):
    if not _is_admin(message):
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        await message.answer("❌ openpyxl kutubxonasi o'rnatilmagan.")
        return

    now = tz_now()
    year, month = now.year, now.month
    employees = get_all_employees(active_only=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{texts.MONTHS_UZ[month]} {year}"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    headers = ["№", "F.I.Sh", "Lavozim", "Sana", "Keldi", "Ketdi", "Ishlagan vaqt"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    row = 2
    idx = 1
    for emp in employees:
        records = get_monthly_attendance(emp["id"], year, month)
        if not records:
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=emp["full_name"])
            ws.cell(row=row, column=3, value=emp["position"])
            ws.cell(row=row, column=4, value="—")
            row += 1
        else:
            for rec in records:
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=emp["full_name"])
                ws.cell(row=row, column=3, value=emp["position"])
                ws.cell(row=row, column=4, value=rec["day"])
                ws.cell(row=row, column=5, value=rec["first_in"][:5] if rec["first_in"] else "—")
                ws.cell(row=row, column=6, value=rec["last_out"][:5] if rec["last_out"] else "—")
                worked = "—"
                if rec["first_in"] and rec["last_out"]:
                    try:
                        ih, im, _ = map(int, rec["first_in"].split(":"))
                        oh, om, _ = map(int, rec["last_out"].split(":"))
                        m = (oh * 60 + om) - (ih * 60 + im)
                        if m > 0:
                            worked = f"{m // 60}s {m % 60}d"
                    except Exception:
                        pass
                ws.cell(row=row, column=7, value=worked)
                row += 1
        idx += 1

    # Ustun kengligi
    widths = [5, 30, 20, 12, 10, 10, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"davomat_{year}_{month:02d}.xlsx"
    await message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=f"📊 <b>{texts.MONTHS_UZ[month]} {year} davomat hisoboti</b>"
    )


# ===== Ishxona sozlamalari =====

@router.message(F.text == texts.BTN_ADMIN_SETTINGS)
async def admin_settings(message: Message, state: FSMContext):
    if not _is_admin(message):
        return
    await state.clear()
    cfg = get_office_config()
    from services.wifi_verify import get_local_ip
    server_ip = get_local_ip()
    await message.answer(
        texts.ADMIN_SETTINGS_VIEW.format(
            server_ip=server_ip,
            start=cfg["work_start"], end=cfg["work_end"]
        ),
        reply_markup=kb.admin_settings_kb()
    )


@router.message(F.text == texts.BTN_SET_HOURS)
async def set_hours_start(message: Message, state: FSMContext):
    if not _is_admin(message):
        return
    await message.answer(texts.ADMIN_SET_WORK_START, reply_markup=kb.cancel_kb())
    await state.set_state(AdminPanel.waiting_work_start)


def _validate_time(text: str) -> bool:
    try:
        h, m = text.strip().split(":")
        h, m = int(h), int(m)
        return 0 <= h <= 23 and 0 <= m <= 59
    except Exception:
        return False


@router.message(AdminPanel.waiting_work_start, F.text)
async def set_work_start(message: Message, state: FSMContext):
    if not _validate_time(message.text):
        await message.answer(texts.ADMIN_TIME_INVALID)
        return
    await state.update_data(work_start=message.text.strip())
    await message.answer(texts.ADMIN_SET_WORK_END, reply_markup=kb.cancel_kb())
    await state.set_state(AdminPanel.waiting_work_end)


@router.message(AdminPanel.waiting_work_end, F.text)
async def set_work_end(message: Message, state: FSMContext):
    if not _validate_time(message.text):
        await message.answer(texts.ADMIN_TIME_INVALID)
        return
    data = await state.get_data()
    set_setting("work_start", data["work_start"])
    set_setting("work_end", message.text.strip())
    await state.clear()
    await message.answer(
        texts.ADMIN_HOURS_DONE.format(start=data["work_start"], end=message.text.strip()),
        reply_markup=kb.admin_settings_kb()
    )

# ===== Davomat tahrirlash =====

@router.message(F.text == texts.BTN_ADMIN_ATT_EDIT)
async def admin_att_edit_start(message: Message, state: FSMContext):
    if not _is_admin(message):
        return
    await state.clear()
    employees = get_all_employees(active_only=True)
    if not employees:
        await message.answer("📋 Xodimlar ro'yxati bo'sh.")
        return
    await message.answer(
        texts.ADMIN_ATT_SELECT_EMPLOYEE,
        reply_markup=kb.employees_inline_kb(employees, "att_edit")
    )


@router.callback_query(F.data.startswith("att_edit:"))
async def admin_att_edit_employee(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    parts = call.data.split(":")
    if parts[1] == "cancel":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    emp_id = int(parts[1])
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    att_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🟢 Keldim qo'shish", callback_data=f"att_add:in:{emp_id}")],
        [InlineKeyboardButton(text="🔴 Ketdim qo'shish", callback_data=f"att_add:out:{emp_id}")],
        [InlineKeyboardButton(text="🗑 Bugungi yozuvlarni tozalash", callback_data=f"att_reset:{emp_id}")],
        [InlineKeyboardButton(text=texts.BTN_CANCEL, callback_data="att_edit:cancel")],
    ])
    await call.message.edit_text(
        texts.ADMIN_ATT_ACTIONS.format(name=emp["full_name"]),
        reply_markup=att_kb
    )
    await call.answer()


@router.callback_query(F.data.startswith("att_add:"))
async def admin_att_add(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    _, check_type, emp_id = call.data.split(":")
    emp = get_employee_by_id(int(emp_id))
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    action_name = "🟢 Keldim" if check_type == "in" else "🔴 Ketdim"
    await state.update_data(att_employee_id=int(emp_id), att_check_type=check_type)
    await call.message.edit_text(
        texts.ADMIN_ATT_ENTER_TIME.format(name=emp["full_name"], action=action_name)
    )
    await state.set_state(AdminPanel.waiting_att_time)
    await call.answer()


@router.message(AdminPanel.waiting_att_time, F.text)
async def admin_att_save_time(message: Message, state: FSMContext):
    if not _validate_time(message.text):
        await message.answer(texts.ADMIN_TIME_INVALID)
        return

    data = await state.get_data()
    emp_id = data["att_employee_id"]
    check_type = data["att_check_type"]
    time_str = message.text.strip()

    emp = get_employee_by_id(emp_id)
    add_manual_attendance(emp_id, check_type, time_str)

    action_name = "🟢 Keldim" if check_type == "in" else "🔴 Ketdim"
    await state.clear()
    await message.answer(
        texts.ADMIN_ATT_SAVED.format(
            name=emp["full_name"], action=action_name, time=time_str
        ),
        reply_markup=_admin_kb(message)
    )


@router.callback_query(F.data.startswith("att_reset:"))
async def admin_att_reset(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    emp_id = int(call.data.split(":")[1])
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    count = delete_today_attendance(emp_id)
    await call.message.edit_text(
        texts.ADMIN_ATT_RESET_DONE.format(name=emp["full_name"], count=count)
    )
    await call.answer("✅ Tozalandi")

# ===== Stavkalarni belgilash =====

@router.message(F.text == texts.BTN_ADMIN_RATES)
async def admin_rates_start(message: Message, state: FSMContext):
    if not _is_admin(message):
        return
    await state.clear()
    employees = get_all_employees(active_only=True)
    if not employees:
        await message.answer("📋 Xodimlar ro'yxati bo'sh.")
        return
    await message.answer(
        texts.ADMIN_RATES_PROMPT,
        reply_markup=kb.employees_inline_kb(employees, "set_rate")
    )


@router.callback_query(F.data.startswith("set_rate:"))
async def admin_rate_choose_employee(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    parts = call.data.split(":")
    if parts[1] == "cancel":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    emp_id = int(parts[1])
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    current = emp["hourly_rate"] if "hourly_rate" in emp.keys() and emp["hourly_rate"] else 0
    await state.update_data(rate_employee_id=emp_id)
    await call.message.edit_text(
        texts.ADMIN_RATE_ENTER.format(name=emp["full_name"], current=current)
    )
    await state.set_state(AdminPanel.waiting_hourly_rate)
    await call.answer()


@router.message(AdminPanel.waiting_hourly_rate, F.text)
async def admin_rate_save(message: Message, state: FSMContext):
    try:
        # Bo'sh joy va vergullarni olib tashlash ("25 000" yoki "25,000" → 25000)
        cleaned = message.text.replace(" ", "").replace(",", "").replace(".", "").strip()
        rate = int(cleaned)
        if rate < 0 or rate > 10_000_000:
            raise ValueError()
    except ValueError:
        await message.answer(texts.ADMIN_RATE_INVALID)
        return

    data = await state.get_data()
    emp_id = data["rate_employee_id"]
    emp = get_employee_by_id(emp_id)
    if not emp:
        await state.clear()
        await message.answer("❌ Xodim topilmadi.", reply_markup=_admin_kb(message))
        return

    set_hourly_rate(emp_id, rate)
    await state.clear()
    await message.answer(
        texts.ADMIN_RATE_SAVED.format(name=emp["full_name"], rate=rate),
        reply_markup=_admin_kb(message)
    )


# ===== Ish haqqi yozuvlarini boshqarish =====

CONFIRM_THRESHOLD = 1_000_000  # 1M so'mdan katta uchun tasdiq so'raladi


async def _send_notification(bot: Bot, telegram_id: int, text: str):
    """Xodimga bildirishnoma yuborish (xatolik chiqsa loglanadi)"""
    try:
        await bot.send_message(telegram_id, text, parse_mode="HTML")
    except Exception as e:
        logger.warning(f"Bildirishnoma yuborib bo'lmadi: tg={telegram_id} xato={e}")


@router.message(F.text == texts.BTN_ADMIN_SALARY)
async def admin_salary_start(message: Message, state: FSMContext):
    if not _is_admin(message):
        return
    await state.clear()
    await message.answer(
        texts.SALARY_ADMIN_MENU,
        reply_markup=kb.salary_admin_menu_kb()
    )


@router.callback_query(F.data == "sal_close")
async def admin_salary_close(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text(texts.CANCELLED)
    await call.answer()


# --- Qo'shish oqimi ---

@router.callback_query(F.data == "sal_add")
async def admin_salary_add_choose_employee(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    # Joriy oy yopiqligini tekshirish
    now = tz_now()
    if is_month_closed(now.year, now.month):
        await call.message.edit_text(
            texts.MONTH_BLOCKED.format(month=texts.MONTHS_UZ[now.month], year=now.year)
        )
        await call.answer()
        return
    employees = get_all_employees(active_only=True)
    if not employees:
        await call.message.edit_text("📋 Xodimlar yo'q.")
        await call.answer()
        return
    await call.message.edit_text(
        texts.SALARY_ADD_CHOOSE_EMPLOYEE,
        reply_markup=kb.salary_employees_kb(employees, "sal_addemp")
    )
    await call.answer()


@router.callback_query(F.data.startswith("sal_addemp:"))
async def admin_salary_choose_type(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    if parts[1] == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return
    emp_id = int(parts[1])
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return
    await state.update_data(sal_emp_id=emp_id, sal_emp_name=emp["full_name"],
                            sal_emp_telegram=emp["telegram_id"])
    await call.message.edit_text(
        texts.SALARY_ADD_CHOOSE_TYPE.format(name=emp["full_name"]),
        reply_markup=kb.salary_types_kb()
    )
    await call.answer()


@router.callback_query(F.data.startswith("sal_type:"))
async def admin_salary_ask_amount(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    if parts[1] == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    entry_type = parts[1]
    type_info = texts.SALARY_TYPES.get(entry_type)
    if not type_info:
        await call.answer("❌ Noma'lum tur", show_alert=True)
        return

    data = await state.get_data()
    await state.update_data(sal_type=entry_type)
    await call.message.edit_text(
        texts.SALARY_ADD_AMOUNT.format(
            emoji=type_info[0], type_name=type_info[1],
            name=data["sal_emp_name"]
        )
    )
    await state.set_state(AdminSalary.add_amount)
    await call.answer()


@router.message(AdminSalary.add_amount, F.text)
async def admin_salary_ask_reason(message: Message, state: FSMContext):
    try:
        cleaned = message.text.replace(" ", "").replace(",", "").replace(".", "").strip()
        amount = int(cleaned)
        if amount <= 0 or amount > 100_000_000:
            raise ValueError()
    except ValueError:
        await message.answer(texts.SALARY_AMOUNT_INVALID)
        return

    data = await state.get_data()
    type_info = texts.SALARY_TYPES[data["sal_type"]]
    await state.update_data(sal_amount=amount)
    await message.answer(
        texts.SALARY_ADD_REASON.format(
            emoji=type_info[0], type_name=type_info[1],
            name=data["sal_emp_name"], amount=amount
        )
    )
    await state.set_state(AdminSalary.add_reason)


@router.message(AdminSalary.add_reason, F.text)
async def admin_salary_handle_reason(message: Message, state: FSMContext, bot: Bot):
    reason = message.text.strip()
    if len(reason) < 3:
        await message.answer(texts.SALARY_REASON_SHORT)
        return

    data = await state.get_data()
    await state.update_data(sal_reason=reason)
    type_info = texts.SALARY_TYPES[data["sal_type"]]

    # Katta summa uchun tasdiq so'rash
    if data["sal_amount"] >= CONFIRM_THRESHOLD:
        await message.answer(
            texts.SALARY_ADD_CONFIRM.format(
                emoji=type_info[0], type_name=type_info[1],
                name=data["sal_emp_name"],
                amount=data["sal_amount"], reason=reason
            ),
            reply_markup=kb.salary_confirm_kb()
        )
        # Hozircha state'da qoldiramiz, callback bilan saqlanadi
    else:
        # To'g'ridan-to'g'ri saqlaymiz
        await _do_save_salary(message, state, bot, data, reason)


@router.callback_query(F.data.startswith("sal_confirm:"))
async def admin_salary_confirm(call: CallbackQuery, state: FSMContext, bot: Bot):
    parts = call.data.split(":")
    if parts[1] != "yes":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    data = await state.get_data()
    if "sal_amount" not in data:
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    # Tasdiqlangan — saqlaymiz
    type_info = texts.SALARY_TYPES[data["sal_type"]]
    me = get_employee_by_telegram_id(call.from_user.id)

    add_salary_entry(
        employee_id=data["sal_emp_id"],
        entry_type=data["sal_type"],
        amount=data["sal_amount"],
        reason=data["sal_reason"],
        created_by=me["id"] if me else 0
    )

    await call.message.edit_text(
        texts.SALARY_ADD_SAVED.format(
            emoji=type_info[0], type_name=type_info[1],
            name=data["sal_emp_name"],
            amount=data["sal_amount"], reason=data["sal_reason"]
        )
    )
    await call.answer("✅ Saqlandi")

    # Xodimga bildirishnoma
    sign = type_info[2]
    notify_text = texts.NOTIFY_SALARY_ADDED.format(
        emoji=type_info[0], type_name=type_info[1],
        sign=sign, amount=data["sal_amount"], reason=data["sal_reason"]
    )
    await _send_notification(bot, data["sal_emp_telegram"], notify_text)

    await state.clear()


async def _do_save_salary(message: Message, state: FSMContext, bot: Bot, data: dict, reason: str):
    """1M dan kichik summa — to'g'ridan-to'g'ri saqlash"""
    type_info = texts.SALARY_TYPES[data["sal_type"]]
    me = get_employee_by_telegram_id(message.from_user.id)

    add_salary_entry(
        employee_id=data["sal_emp_id"],
        entry_type=data["sal_type"],
        amount=data["sal_amount"],
        reason=reason,
        created_by=me["id"] if me else 0
    )

    await message.answer(
        texts.SALARY_ADD_SAVED.format(
            emoji=type_info[0], type_name=type_info[1],
            name=data["sal_emp_name"],
            amount=data["sal_amount"], reason=reason
        ),
        reply_markup=_admin_kb(message)
    )

    # Bildirishnoma
    sign = type_info[2]
    notify_text = texts.NOTIFY_SALARY_ADDED.format(
        emoji=type_info[0], type_name=type_info[1],
        sign=sign, amount=data["sal_amount"], reason=reason
    )
    await _send_notification(bot, data["sal_emp_telegram"], notify_text)

    await state.clear()


# --- Bekor qilish oqimi ---

@router.callback_query(F.data == "sal_cancel")
async def admin_salary_cancel_choose_emp(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    # Joriy oy yopiqligini tekshirish
    now = tz_now()
    if is_month_closed(now.year, now.month):
        await call.message.edit_text(
            texts.MONTH_BLOCKED.format(month=texts.MONTHS_UZ[now.month], year=now.year)
        )
        await call.answer()
        return
    employees = get_all_employees(active_only=True)
    if not employees:
        await call.message.edit_text("📋 Xodimlar yo'q.")
        await call.answer()
        return
    await call.message.edit_text(
        texts.SALARY_CANCEL_CHOOSE_EMPLOYEE,
        reply_markup=kb.salary_employees_kb(employees, "sal_cancelemp")
    )
    await call.answer()


@router.callback_query(F.data.startswith("sal_cancelemp:"))
async def admin_salary_cancel_choose_entry(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    if parts[1] == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return
    emp_id = int(parts[1])
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    now = tz_now()
    entries = get_active_salary_entries(emp_id, now.year, now.month)
    if not entries:
        await call.message.edit_text(
            texts.SALARY_CANCEL_NO_ENTRIES.format(name=emp["full_name"])
        )
        await call.answer()
        return

    await state.update_data(sal_emp_id=emp_id, sal_emp_name=emp["full_name"],
                            sal_emp_telegram=emp["telegram_id"])
    await call.message.edit_text(
        texts.SALARY_CANCEL_CHOOSE_ENTRY.format(name=emp["full_name"]),
        reply_markup=kb.salary_entries_kb(entries, "sal_canc")
    )
    await call.answer()


@router.callback_query(F.data.startswith("sal_canc:"))
async def admin_salary_cancel_ask_reason(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    if parts[1] == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return
    entry_id = int(parts[1])
    entry = get_salary_entry(entry_id)
    if not entry:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    type_info = texts.SALARY_TYPES.get(entry["entry_type"], ("📋", "?", ""))
    await state.update_data(sal_cancel_entry_id=entry_id,
                            sal_cancel_type=entry["entry_type"],
                            sal_cancel_amount=entry["amount"],
                            sal_cancel_reason_orig=entry["reason"] or "—")
    await call.message.edit_text(
        texts.SALARY_CANCEL_REASON.format(
            emoji=type_info[0], type_name=type_info[1],
            amount=entry["amount"], reason=entry["reason"] or "—"
        )
    )
    await state.set_state(AdminSalary.cancel_reason)
    await call.answer()


@router.message(AdminSalary.cancel_reason, F.text)
async def admin_salary_cancel_save(message: Message, state: FSMContext, bot: Bot):
    cancel_reason = message.text.strip()
    if len(cancel_reason) < 3:
        await message.answer(texts.SALARY_REASON_SHORT)
        return

    data = await state.get_data()
    me = get_employee_by_telegram_id(message.from_user.id)
    cancel_salary_entry(
        entry_id=data["sal_cancel_entry_id"],
        cancelled_by=me["id"] if me else 0,
        cancel_reason=cancel_reason
    )

    type_info = texts.SALARY_TYPES.get(data["sal_cancel_type"], ("📋", "?", ""))
    await message.answer(
        texts.SALARY_CANCEL_DONE.format(
            emoji=type_info[0], type_name=type_info[1],
            amount=data["sal_cancel_amount"], cancel_reason=cancel_reason
        ),
        reply_markup=_admin_kb(message)
    )

    # Bildirishnoma
    notify_text = texts.NOTIFY_SALARY_CANCELLED.format(
        emoji=type_info[0], type_name=type_info[1],
        amount=data["sal_cancel_amount"], cancel_reason=cancel_reason
    )
    await _send_notification(bot, data["sal_emp_telegram"], notify_text)

    await state.clear()


# ===== Phase 3: Excel hisobot =====

def _generate_salary_excel(year: int, month: int) -> bytes:
    """Excel fayl: barcha xodimlarning oylik ish haqqi"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Sheet 1: Xulosa
    ws = wb.active
    ws.title = f"{texts.MONTHS_UZ[month]} {year}"

    title = ws.cell(row=1, column=1,
                    value=f"📊 Ish haqqi hisoboti — {texts.MONTHS_UZ[month]} {year}")
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.row_dimensions[1].height = 25

    headers = ["№", "Xodim", "Lavozim", "Soat", "Stavka (so'm/s)",
               "Asosiy ish haqqi", "Avans", "Jarima", "Mukofot", "Bonus",
               "Mahsulot", "JAMI (so'm)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[3].height = 35

    summary = get_all_employees_salary_summary(year, month)
    grand_total = 0
    for idx, item in enumerate(summary, 1):
        emp = item["employee"]
        row = idx + 3
        hours_str = f"{item['minutes'] // 60}s {item['minutes'] % 60}d"
        values = [idx, emp["full_name"], emp["position"], hours_str,
                  item["rate"], item["base"],
                  -item["totals"]["avans"], -item["totals"]["jarima"],
                  item["totals"]["mukofot"], item["totals"]["bonus"],
                  -item["totals"]["mahsulot"], item["total"]]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border
            if col == 12:
                c.font = Font(bold=True)
                if v > 0:
                    c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            if col in (5, 6, 7, 8, 9, 10, 11, 12):
                c.number_format = '#,##0'
        grand_total += item["total"]

    # Pastdagi jami qator
    total_row = len(summary) + 4
    c = ws.cell(row=total_row, column=11, value="JAMI:")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="right")
    c = ws.cell(row=total_row, column=12, value=grand_total)
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
    c.number_format = '#,##0'

    # Ustun kengligi
    widths = [4, 25, 15, 10, 14, 16, 12, 12, 12, 12, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Sheet 2: Yozuvlar tafsiloti
    ws2 = wb.create_sheet(title="Yozuvlar")
    ws2.cell(row=1, column=1, value="Sana").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Xodim").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Kategoriya").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="Summa").font = Font(bold=True)
    ws2.cell(row=1, column=5, value="Sabab").font = Font(bold=True)
    ws2.cell(row=1, column=6, value="Qo'shgan").font = Font(bold=True)
    ws2.cell(row=1, column=7, value="Holat").font = Font(bold=True)
    for col in range(1, 8):
        c = ws2.cell(row=1, column=col)
        c.fill = header_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")

    audit = get_audit_entries(year, month, limit=1000)
    for idx, e in enumerate(audit, 2):
        try:
            d = to_local(e["created_at"])
            date_str = d.strftime("%d.%m.%Y %H:%M")
        except Exception:
            date_str = e["created_at"]
        type_info = texts.SALARY_TYPES.get(e["entry_type"], ("", "?", "+"))
        ws2.cell(row=idx, column=1, value=date_str)
        ws2.cell(row=idx, column=2, value=e["employee_name"] or "—")
        ws2.cell(row=idx, column=3, value=f"{type_info[0]} {type_info[1]}")
        c = ws2.cell(row=idx, column=4, value=e["amount"])
        c.number_format = '#,##0'
        ws2.cell(row=idx, column=5, value=e["reason"] or "—")
        ws2.cell(row=idx, column=6, value=e["creator_name"] or "—")
        if e["cancelled"]:
            status = f"❌ Bekor ({e['canceller_name'] or '—'}): {e['cancel_reason'] or '—'}"
        else:
            status = "✅ Faol"
        ws2.cell(row=idx, column=7, value=status)

    ws2_widths = [18, 25, 20, 14, 35, 20, 40]
    for i, w in enumerate(ws2_widths, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.callback_query(F.data == "sal_report")
async def admin_salary_report(call: CallbackQuery, state: FSMContext, bot: Bot):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    await call.message.edit_text(texts.SAL_REPORT_GENERATING)
    await call.answer()

    now = tz_now()
    year, month = now.year, now.month
    file_bytes = await asyncio.to_thread(_generate_salary_excel, year, month)

    await call.message.answer_document(
        BufferedInputFile(file_bytes, filename=f"ish_haqqi_{year}_{month:02d}.xlsx"),
        caption=texts.SAL_REPORT_DONE.format(month=texts.MONTHS_UZ[month], year=year),
        reply_markup=_admin_kb(call)
    )


# ===== Phase 3: Audit tarixi =====

@router.callback_query(F.data == "sal_audit")
async def admin_salary_audit(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    now = tz_now()
    entries = get_audit_entries(now.year, now.month, limit=30)
    if not entries:
        await call.message.edit_text(texts.AUDIT_EMPTY)
        await call.answer()
        return

    text = texts.AUDIT_HEADER.format(
        month=texts.MONTHS_UZ[now.month], year=now.year, limit=len(entries)
    )
    for e in entries:
        type_info = texts.SALARY_TYPES.get(e["entry_type"], ("📋", "?", "+"))
        try:
            d = to_local(e["created_at"])
            date_str = d.strftime("%d.%m %H:%M")
        except Exception:
            date_str = "—"
        text += texts.AUDIT_LINE.format(
            date=date_str,
            emoji=type_info[0], type_name=type_info[1],
            sign=type_info[2], amount=e["amount"],
            employee=e["employee_name"] or "—",
            creator=e["creator_name"] or "—",
            reason=e["reason"] or "—"
        )
        if e["cancelled"]:
            text += texts.AUDIT_CANCELLED.format(
                canceller=e["canceller_name"] or "—",
                cancel_reason=e["cancel_reason"] or "—"
            )

    # Telegram message limit
    if len(text) > 4000:
        text = text[:3900] + "\n\n<i>... (qisqartirildi)</i>"

    await call.message.edit_text(text)
    await call.answer()


# ===== Phase 3: Oy yopish/ochish =====

@router.callback_query(F.data == "sal_close_month")
async def admin_close_month_prompt(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    now = tz_now()
    if is_month_closed(now.year, now.month):
        # Allaqachon yopilgan — qayta ochish taklif qilamiz
        await call.message.edit_text(
            f"🔒 <b>{texts.MONTHS_UZ[now.month]} {now.year}</b> oyi allaqachon yopilgan.\n\n"
            f"Qayta ochaymi?",
            reply_markup=kb.month_reopen_confirm_kb()
        )
    else:
        await call.message.edit_text(
            texts.MONTH_CLOSE_CONFIRM.format(
                month=texts.MONTHS_UZ[now.month], year=now.year
            ),
            reply_markup=kb.month_close_confirm_kb()
        )
    await call.answer()


@router.callback_query(F.data == "sal_cm_yes")
async def admin_close_month_do(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    now = tz_now()
    close_month(now.year, now.month, me["id"])
    await call.message.edit_text(
        texts.MONTH_CLOSE_DONE.format(
            month=texts.MONTHS_UZ[now.month], year=now.year
        )
    )
    await call.answer("🔒 Yopildi")


@router.callback_query(F.data == "sal_cm_reopen")
async def admin_reopen_month(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or not me["is_admin"]:
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    now = tz_now()
    reopen_month(now.year, now.month)
    await call.message.edit_text(
        texts.MONTH_REOPEN_DONE.format(
            month=texts.MONTHS_UZ[now.month], year=now.year
        )
    )
    await call.answer("🔓 Ochildi")


@router.callback_query(F.data == "sal_cm_no")
async def admin_close_month_no(call: CallbackQuery, state: FSMContext):
    await call.message.edit_text(texts.CANCELLED)
    await call.answer()


# ===== Phase 2: Admin vazifa berish =====

@router.message(F.text == texts.BTN_ADMIN_TASKS)
async def task_pick_employee(message: Message, state: FSMContext):
    """Admin yoki Boss "Vazifa berish" tugmasini bosdi — xodim tanlash."""
    from roles import can_assign_tasks
    me = get_employee_by_telegram_id(message.from_user.id)
    if not me or not can_assign_tasks(me):
        await message.answer(texts.NO_PERMISSION)
        return

    await state.clear()
    employees = get_all_employees(active_only=True)
    # O'zini ham ro'yxatdan chiqarmaymiz — admin o'ziga ham vazifa qo'sha oladi
    if not employees:
        await message.answer("❌ Xodimlar yo'q.")
        return

    await message.answer(
        texts.ADMIN_TASK_PICK_EMP,
        reply_markup=kb.salary_employees_kb(employees, prefix="task_emp")
    )
    await state.set_state(TaskCreate.choosing_employee)


@router.callback_query(TaskCreate.choosing_employee, F.data.startswith("task_emp:"))
async def task_employee_chosen(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    if parts[1] == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    try:
        emp_id = int(parts[1])
    except ValueError:
        await call.answer("❌ Xato", show_alert=True)
        return
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    await state.update_data(task_emp_id=emp_id,
                            task_emp_name=emp["full_name"],
                            task_emp_telegram=emp["telegram_id"])
    await call.message.edit_text(
        texts.ADMIN_TASK_ASK_TITLE.format(name=emp["full_name"])
    )
    await state.set_state(TaskCreate.entering_title)
    await call.answer()


@router.message(TaskCreate.entering_title, F.text)
async def task_title_handler(message: Message, state: FSMContext):
    title = message.text.strip()
    if message.text == texts.BTN_CANCEL:
        await state.clear()
        await message.answer(texts.CANCELLED,
                             reply_markup=_admin_kb(message))
        return
    if len(title) < 3:
        await message.answer(texts.ADMIN_TASK_TITLE_SHORT)
        return

    await state.update_data(task_title=title)
    await message.answer(
        texts.ADMIN_TASK_ASK_DESC,
        reply_markup=kb.task_description_skip_kb()
    )
    await state.set_state(TaskCreate.entering_description)


@router.message(TaskCreate.entering_description, F.text)
async def task_description_handler(message: Message, state: FSMContext):
    if message.text == texts.BTN_CANCEL:
        await state.clear()
        await message.answer(texts.CANCELLED,
                             reply_markup=_admin_kb(message))
        return

    if message.text == texts.BTN_TASK_SKIP_DESCRIPTION:
        desc = None
    else:
        desc = message.text.strip()
        if len(desc) > 500:
            desc = desc[:500]

    await state.update_data(task_description=desc)
    await message.answer(
        texts.ADMIN_TASK_ASK_DEADLINE,
        reply_markup=kb.task_deadline_skip_kb()
    )
    await state.set_state(TaskCreate.entering_deadline)


def _parse_deadline(text: str):
    """DD.MM yoki DD.MM HH:MM kiritishni UTC ISO string'ga aylantiradi (joriy yil)."""
    text = text.strip()
    now_local = tz_now()
    # Format variantlari
    formats = ("%d.%m %H:%M", "%d.%m")
    parsed = None
    for f in formats:
        try:
            t = datetime.strptime(text, f).replace(year=now_local.year)
            if f == "%d.%m":
                # Standart muddat: kun oxiri 23:59
                t = t.replace(hour=23, minute=59)
            parsed = t
            break
        except ValueError:
            continue
    if parsed is None:
        return None
    # Local Tashkent -> UTC saqlash
    from tzutil import OFFSET
    return (parsed - OFFSET).isoformat()


@router.message(TaskCreate.entering_deadline, F.text)
async def task_deadline_handler(message: Message, state: FSMContext, bot: Bot):
    if message.text == texts.BTN_CANCEL:
        await state.clear()
        await message.answer(texts.CANCELLED,
                             reply_markup=_admin_kb(message))
        return

    if message.text == texts.BTN_TASK_SKIP_DEADLINE:
        deadline_iso = None
    else:
        deadline_iso = _parse_deadline(message.text)
        if deadline_iso is None:
            await message.answer(texts.ADMIN_TASK_DEADLINE_INVALID)
            return

    data = await state.get_data()
    me = get_employee_by_telegram_id(message.from_user.id)

    task_id = create_task(
        title=data["task_title"],
        description=data.get("task_description"),
        assigned_to=data["task_emp_id"],
        assigned_by=me["id"],
        deadline=deadline_iso,
    )

    deadline_str = ""
    if deadline_iso:
        deadline_str = "\n⏰ Muddat: <i>{}</i>".format(
            to_local(deadline_iso).strftime("%d.%m %H:%M")
        )
    desc_str = ""
    if data.get("task_description"):
        desc_str = "\n💬 Izoh: <i>{}</i>".format(data["task_description"])

    await message.answer(
        texts.ADMIN_TASK_SAVED.format(
            name=data["task_emp_name"],
            title=data["task_title"],
            deadline=deadline_str,
            desc=desc_str,
        ),
        reply_markup=_admin_kb(message)
    )

    # Xodimga bildirishnoma
    notify_text = texts.ADMIN_TASK_NOTIFY_EMP.format(
        title=data["task_title"],
        by=me["full_name"],
        deadline=deadline_str,
        desc=desc_str,
    )
    await _send_notification(bot, data["task_emp_telegram"], notify_text)

    await state.clear()


# ===== Phase 3A: Bosh Admin tomonidan Boss tayinlash =====

@router.message(F.text == texts.BTN_ADMIN_BOSS_ASSIGN)
async def admin_boss_pick(message: Message, state: FSMContext):
    """Bosh Admin Boss tayinlash tugmasini bosdi."""
    me = get_employee_by_telegram_id(message.from_user.id)
    if not me or me["role"] != "bosh_admin":
        await message.answer(texts.ADMIN_BOSS_ONLY_BOSH)
        return

    await state.clear()
    current_boss = get_boss()
    current_line = (
        texts.ADMIN_BOSS_CURRENT.format(name=current_boss["full_name"])
        if current_boss else texts.ADMIN_BOSS_NONE_YET
    )

    # Adminlar va Bosh Admin'ni ham ro'yxatdan chiqarmaymiz —
    # kerak bo'lsa adminni Boss qilish mumkin. Faqat o'zini chiqarib qo'yamiz.
    employees = [e for e in get_all_employees(active_only=True)
                 if e["id"] != me["id"]]
    if not employees:
        await message.answer("❌ Tanlash uchun xodim yo'q.",
                             reply_markup=_admin_kb(message))
        return

    await message.answer(
        texts.ADMIN_BOSS_PICK.format(current=current_line),
        reply_markup=kb.salary_employees_kb(employees, prefix="boss_pick")
    )


@router.callback_query(F.data.startswith("boss_pick:"))
async def admin_boss_pick_callback(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or me["role"] != "bosh_admin":
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    parts = call.data.split(":")
    if parts[1] == "cancel":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    try:
        emp_id = int(parts[1])
    except ValueError:
        await call.answer("❌ Xato", show_alert=True)
        return
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return
    if emp["role"] == "bosh_admin":
        await call.answer("❌ Bosh Admin Boss bo'la olmaydi.", show_alert=True)
        return

    current_boss = get_boss()
    warning = ""
    if current_boss and current_boss["id"] != emp["id"]:
        warning = texts.ADMIN_BOSS_WARNING_REPLACE.format(old=current_boss["full_name"])

    await call.message.edit_text(
        texts.ADMIN_BOSS_CONFIRM.format(name=emp["full_name"], warning=warning),
        reply_markup=kb.assign_boss_confirm_kb(emp_id)
    )
    await call.answer()


@router.callback_query(F.data.startswith("boss_set:"))
async def admin_boss_set_confirm(call: CallbackQuery, bot: Bot):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not me or me["role"] != "bosh_admin":
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    parts = call.data.split(":")
    if parts[1] != "yes":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    try:
        emp_id = int(parts[2])
    except (IndexError, ValueError):
        await call.answer("❌ Xato", show_alert=True)
        return
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Topilmadi", show_alert=True)
        return

    # Eski Bossni xodim qilamiz
    current_boss = get_boss()
    if current_boss and current_boss["id"] != emp_id:
        set_role(current_boss["id"], "employee")
        await _send_notification(bot, current_boss["telegram_id"],
                                 texts.BOSS_NOTIFY_REMOVED)

    # Yangi Boss
    set_role(emp_id, "boss")
    await call.message.edit_text(texts.ADMIN_BOSS_DONE.format(name=emp["full_name"]))
    await call.answer("✅")
    await _send_notification(bot, emp["telegram_id"], texts.BOSS_NOTIFY_ASSIGNED)
