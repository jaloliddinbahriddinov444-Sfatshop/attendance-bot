"""Admin: Xodimlar ma'lumotlari — lavozim bo'yicha guruhlash + to'liq profil."""
import io
import re
import asyncio
import calendar
import logging
from datetime import date as _date
from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, BufferedInputFile

import texts
import keyboards as kb
from tzutil import now as tz_now, to_local
from database import (
    get_employee_by_telegram_id, get_all_employees, get_employee_by_id,
    get_all_positions, get_position,
    get_monthly_attendance, get_monthly_worked_minutes, get_monthly_base_salary,
    get_active_salary_entries, get_salary_totals_by_type,
    get_open_tasks_with_skips,
)

logger = logging.getLogger(__name__)
router = Router()

WEEKDAYS = ["Du", "Se", "Cho", "Pa", "Ju", "Sha", "Ya"]
WEEKDAYS_FULL = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]

_UNSAFE_CHARS = re.compile(r'[\[\]:*?/\\]')


def _safe_sheet_name(name: str, used: set) -> str:
    clean = _UNSAFE_CHARS.sub("", name)[:31] or "Xodim"
    base, idx = clean, 2
    while clean in used:
        clean = base[: 29 - len(str(idx))] + f"_{idx}"
        idx += 1
    used.add(clean)
    return clean


def _fill_emp_sheet(ws, emp, year: int, month: int) -> None:
    """Bitta worksheet'ni xodim ma'lumotlari, davomat va ish haqqi bilan to'ldiradi."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Side(border_style="thin", color="AAAAAA")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    blue_fill = PatternFill("solid", fgColor="4472C4")
    light_fill = PatternFill("solid", fgColor="EBF3FB")
    absent_fill = PatternFill("solid", fgColor="FFE0E0")
    net_fill = PatternFill("solid", fgColor="2E5C8A")
    base_fill = PatternFill("solid", fgColor="E2EFDA")

    role_map = {"bosh_admin": "Bosh Admin", "boss": "Boss", "admin": "Admin"}
    role_label = role_map.get(
        emp["role"] if "role" in emp.keys() else "", "Xodim"
    )

    # === Xodim ma'lumotlari (1–5-qatorlar) ===
    info = [
        ("Xodim:", emp["full_name"]),
        ("Telefon:", emp["phone"]),
        ("Lavozim:", emp["position"]),
        ("Rol:", role_label),
        ("Hisobot davri:", f"{texts.MONTHS_UZ[month]} {year}"),
    ]
    for r, (label, val) in enumerate(info, 1):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True)
        lc.fill = light_fill
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

    # === Davomat jadvali sarlavhasi (7-qator) ===
    header_row = 7
    headers = ["Sana", "Hafta kuni", "Kirish", "Chiqish", "Ishlagan vaqt", "Kunlik ish haqqi"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = blue_fill
        c.alignment = Alignment(horizontal="center")
        c.border = brd

    # === Stavka logikasi ===
    daily_rate = (emp["daily_rate"] if "daily_rate" in emp.keys() else 0) or 0
    position_id = emp["position_id"] if "position_id" in emp.keys() else None
    hourly_rate = (emp["hourly_rate"] if "hourly_rate" in emp.keys() else 0) or 0

    standard_minutes = None
    if daily_rate and position_id:
        pos = get_position(position_id)
        if pos:
            standard_minutes = pos["work_hours"] * 60

    # === Davomat ma'lumotlari ===
    records = get_monthly_attendance(emp["id"], year, month)
    records_by_day = {r["day"]: r for r in records}

    _, days_in_month = calendar.monthrange(year, month)
    cur_row = header_row + 1
    total_days = 0
    total_minutes = 0
    col_base_sum = 0  # per-day column sum (for reference only)

    for day_num in range(1, days_in_month + 1):
        d = _date(year, month, day_num)
        date_str = d.strftime("%Y-%m-%d")
        date_label = d.strftime("%d.%m.%Y")
        weekday_label = WEEKDAYS_FULL[d.weekday()]

        rec = records_by_day.get(date_str)
        if rec and rec["first_in"]:
            fi_str = rec["first_in"][:5]
            lo_str = rec["last_out"][:5] if rec["last_out"] else "—"

            worked_mins = 0
            if rec["last_out"]:
                try:
                    ih, im, _ = map(int, rec["first_in"].split(":"))
                    oh, om, _ = map(int, rec["last_out"].split(":"))
                    worked_mins = max(0, (oh * 60 + om) - (ih * 60 + im))
                except Exception:
                    pass

            worked_str = f"{worked_mins // 60}s {worked_mins % 60}d" if worked_mins else "—"

            if worked_mins:
                if standard_minutes and daily_rate:
                    capped = min(worked_mins, 12 * 60)
                    day_earn = int(capped / standard_minutes * daily_rate)
                elif hourly_rate:
                    day_earn = int((worked_mins / 60.0) * hourly_rate)
                else:
                    day_earn = 0
            else:
                day_earn = 0

            total_days += 1
            total_minutes += worked_mins
            col_base_sum += day_earn
            row_fill = None
        else:
            fi_str = "—"
            lo_str = "—"
            worked_str = "kelmagan"
            day_earn = 0
            row_fill = absent_fill

        row_vals = [date_label, weekday_label, fi_str, lo_str, worked_str,
                    day_earn if day_earn else ""]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=cur_row, column=col, value=val)
            c.border = brd
            if row_fill:
                c.fill = row_fill
            if col == 6 and isinstance(val, int):
                c.number_format = "#,##0"
        cur_row += 1

    # === Xulosa bloki ===
    # Authoritative figures from DB — matches in-app view
    base = get_monthly_base_salary(emp["id"], year, month)
    entries = get_active_salary_entries(emp["id"], year, month)
    totals = get_salary_totals_by_type(emp["id"], year, month)
    net = (base - totals["avans"] - totals["jarima"]
           + totals["mukofot"] + totals["bonus"] - totals["mahsulot"])

    gap_row = cur_row + 1
    summary = [
        ("📅 Jami ish kunlari:", f"{total_days} kun"),
        ("⏱ Jami ishlangan vaqt:", f"{total_minutes // 60} soat {total_minutes % 60} daqiqa"),
        ("💵 Asosiy ish haqqi:", base),
    ]
    for e in entries:
        info_t = texts.SALARY_TYPES.get(e["entry_type"], ("📋", "?", ""))
        emoji, type_name, sign = info_t
        summary.append((f"  {emoji} {type_name}:", f"{sign}{e['amount']:,} so'm  {e['reason'] or ''}".strip()))
    summary.append(("💰 QOLDIQ ISH HAQQI:", net))

    for i, (label, value) in enumerate(summary):
        r = gap_row + i
        is_net = label.startswith("💰")
        is_base = label.startswith("💵")
        fill = net_fill if is_net else (base_fill if is_base else PatternFill())
        txt_color = "FFFFFF" if is_net else "000000"

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, color=txt_color)
        lc.fill = fill

        vc = ws.cell(row=r, column=2, value=value)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        vc.font = Font(bold=True, color=txt_color, size=13 if is_net else 11)
        vc.fill = fill
        if isinstance(value, int):
            vc.number_format = "#,##0"

    # === Ustun kengliklarini sozlash ===
    for i, w in enumerate([14, 14, 10, 10, 16, 18], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


def _build_emp_excel(emp_id: int, year: int, month: int) -> bytes:
    """Bitta xodim uchun Excel fayl (bytes)."""
    from openpyxl import Workbook
    emp = get_employee_by_id(emp_id)
    if not emp:
        return b""
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(emp["full_name"], set())
    _fill_emp_sheet(ws, emp, year, month)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_all_emp_excel(year: int, month: int) -> bytes:
    """Barcha faol xodimlar — har biri alohida sheet."""
    from openpyxl import Workbook
    employees = [e for e in get_all_employees(active_only=True) if e["telegram_id"] > 0]
    wb = Workbook()
    used_names: set = set()
    first = True
    for emp in employees:
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = _safe_sheet_name(emp["full_name"], used_names)
        _fill_emp_sheet(ws, emp, year, month)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _is_admin(uid: int) -> bool:
    emp = get_employee_by_telegram_id(uid)
    return bool(emp and emp["is_admin"])


def _fmt_time(ts: str, fmt="%H:%M") -> str:
    try:
        return to_local(ts).strftime(fmt)
    except Exception:
        return "—"


def _build_full_profile(emp_id: int) -> str:
    """Xodimning to'liq profili — admin uchun."""
    emp = get_employee_by_id(emp_id)
    if not emp:
        return "❌ Xodim topilmadi."

    now = tz_now()
    year, month = now.year, now.month

    # Rol belgisi
    role = emp["role"] if "role" in emp.keys() else "employee"
    role_badge = {
        "bosh_admin": " 👑",
        "boss": " 🏆",
        "admin": " 🛠",
    }.get(role, "")

    # Karta
    card_num = emp["card_number"] if "card_number" in emp.keys() else ""
    card_holder = emp["card_holder_name"] if "card_holder_name" in emp.keys() else ""
    card = texts.format_card(card_num, card_holder)

    # Ro'yxatdan sana
    try:
        reg = to_local(emp["registered_at"]).strftime("%d.%m.%Y")
    except Exception:
        reg = "—"

    # Lavozim va stavka
    pos_id = emp["position_id"] if "position_id" in emp.keys() else None
    daily_rate = emp["daily_rate"] if "daily_rate" in emp.keys() else 0
    pos = get_position(pos_id) if pos_id else None
    pos_name = pos["name"] if pos else emp["position"]

    out = texts.EMP_FULL_PROFILE.format(
        name=emp["full_name"],
        position=pos_name,
        role_badge=role_badge,
        phone=emp["phone"],
        card=card,
        registered=reg,
    )
    if daily_rate and pos:
        out += texts.EMP_FULL_POSITION_RATE.format(
            rate=daily_rate, hours=pos["work_hours"]
        )

    # Davomat — bu oy
    records = get_monthly_attendance(emp_id, year, month)
    out += texts.EMP_FULL_MONTH_HEADER.format(
        month=texts.MONTHS_UZ[month], year=year
    )
    total_minutes = 0
    for _r in records:
        if _r["first_in"] and _r["last_out"]:
            try:
                _ih,_im,_=map(int,_r["first_in"].split(":"))
                _oh,_om,_=map(int,_r["last_out"].split(":"))
                _m=(_oh*60+_om)-(_ih*60+_im)
                if _m>0: total_minutes+=_m
            except Exception: pass
    days_shown = 0
    if records:
        for rec in records[-15:]:  # Oxirgi 15 kun
            day_str = rec["day"]
            try:
                from datetime import date as _date
                d = _date.fromisoformat(day_str)
                wd = WEEKDAYS[d.weekday()]
                date_label = d.strftime("%d.%m")
            except Exception:
                wd = "—"
                date_label = day_str

            fi = rec["first_in"]
            lo = rec["last_out"]

            if fi and lo:
                try:
                    ih, im, _ = map(int, fi.split(":"))
                    oh, om, _ = map(int, lo.split(":"))
                    mins = (oh * 60 + om) - (ih * 60 + im)
                    if mins > 0:
                        worked_str = f"{mins//60}s {mins%60}d"
                    else:
                        worked_str = "—"
                except Exception:
                    worked_str = "—"
                out += texts.EMP_FULL_DAY_LINE.format(
                    date=date_label, weekday=wd,
                    inn=fi[:5], out=lo[:5], worked=worked_str
                )
            elif fi:
                out += texts.EMP_FULL_DAY_NO_OUT.format(
                    date=date_label, weekday=wd, inn=fi[:5]
                )
            else:
                out += texts.EMP_FULL_DAY_ABSENT.format(
                    date=date_label, weekday=wd
                )
            days_shown += 1
    else:
        out += "  <i>Bu oyda davomat yo'q</i>\n"

    out += texts.EMP_FULL_MONTH_TOTAL.format(
        hours=total_minutes // 60, mins=total_minutes % 60
    )

    # Ish haqqi — bu oy
    base = get_monthly_base_salary(emp_id, year, month)
    totals = get_salary_totals_by_type(emp_id, year, month)
    total_sal = (base - totals["avans"] - totals["jarima"]
                 + totals["mukofot"] + totals["bonus"] - totals["mahsulot"])

    out += texts.EMP_FULL_SALARY_HEADER.format(
        month=texts.MONTHS_UZ[month], year=year
    )
    out += texts.EMP_FULL_SALARY_BASE.format(base=base)
    entries = get_active_salary_entries(emp_id, year, month)
    for e in entries:
        info = texts.SALARY_TYPES.get(e["entry_type"], ("📋", "?", ""))
        emoji, type_name, sign = info
        out += texts.EMP_FULL_SALARY_LINE.format(
            emoji=emoji, type=type_name, sign=sign,
            amount=e["amount"], reason=e["reason"] or "—"
        )
    out += texts.EMP_FULL_SALARY_TOTAL.format(total=total_sal)

    # Vazifalar
    open_tasks = get_open_tasks_with_skips(emp_id)
    if open_tasks:
        out += texts.EMP_FULL_TASKS_HEADER
        for t in open_tasks:
            skip_frag = ""
            if t["skip_count"]:
                skip_frag = texts.EMP_DETAIL_TASK_SKIPS.format(count=t["skip_count"])
            out += texts.EMP_FULL_TASK_LINE.format(
                title=t["title"],
                by=t["assigned_by_name"] or "—",
                skips=skip_frag,
            )

    return out


# ===== Handlerlar =====

@router.message(F.text == texts.BTN_ADMIN_LIST)
async def emp_data_start(message: Message, state: FSMContext):
    if not _is_admin(message.from_user.id):
        return
    await state.clear()
    positions = get_all_positions()

    # Lavozim belgilanmagan xodimlar soni
    all_emps = [e for e in get_all_employees(active_only=True) if e["telegram_id"] > 0]
    assigned_ids = set()
    for pos in positions:
        for emp in all_emps:
            if emp["position_id"] == pos["id"]:
                assigned_ids.add(emp["id"])
    unassigned = len([e for e in all_emps if e["id"] not in assigned_ids])

    await message.answer(
        texts.EMP_DATA_PICK_POS,
        reply_markup=kb.emp_positions_kb(positions, unassigned)
    )


@router.callback_query(F.data == "empdata_back")
async def emp_data_back(call: CallbackQuery):
    positions = get_all_positions()
    all_emps = [e for e in get_all_employees(active_only=True) if e["telegram_id"] > 0]
    assigned_ids = set()
    for pos in positions:
        for emp in all_emps:
            if emp["position_id"] == pos["id"]:
                assigned_ids.add(emp["id"])
    unassigned = len([e for e in all_emps if e["id"] not in assigned_ids])
    await call.message.edit_text(
        texts.EMP_DATA_PICK_POS,
        reply_markup=kb.emp_positions_kb(positions, unassigned)
    )
    await call.answer()


@router.callback_query(F.data.startswith("empdata_pos:"))
async def emp_data_pick_pos(call: CallbackQuery):
    if not _is_admin(call.from_user.id):
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    val = call.data.split(":")[1]
    if val == "cancel":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    pos_id = int(val)
    all_emps = [e for e in get_all_employees(active_only=True) if e["telegram_id"] > 0]

    if pos_id == 0:
        # Lavozim belgilanmagan
        assigned_pos_ids = {e["position_id"] for e in all_emps if e["position_id"]}
        emps = [e for e in all_emps if not e["position_id"]]
        pos_label = "Lavozim belgilanmagan"
    else:
        pos = get_position(pos_id)
        if not pos:
            await call.answer("Topilmadi", show_alert=True)
            return
        emps = [e for e in all_emps if e["position_id"] == pos_id]
        pos_label = pos["name"]

    if not emps:
        await call.answer(texts.EMP_DATA_NO_EMPS, show_alert=True)
        return

    await call.message.edit_text(
        texts.EMP_DATA_PICK_EMP.format(pos_name=pos_label, count=len(emps)),
        reply_markup=kb.emp_in_position_kb(emps, pos_id)
    )
    await call.answer()


@router.callback_query(F.data.startswith("empdata_emp:"))
async def emp_data_detail(call: CallbackQuery):
    if not _is_admin(call.from_user.id):
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    emp_id = int(call.data.split(":")[1])
    profile_text = _build_full_profile(emp_id)

    # Telegram xabar limiti: 4096 belgi
    if len(profile_text) > 4000:
        profile_text = profile_text[:4000] + "\n\n<i>... (qisqartirildi)</i>"

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    back_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Excel hisobot", callback_data=f"empdata_excel:{emp_id}")],
        [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"empdata_back_emp:{emp_id}")],
    ])

    await call.message.edit_text(profile_text, reply_markup=back_kb)
    await call.answer()


@router.callback_query(F.data.startswith("empdata_back_emp:"))
async def emp_data_back_to_pos(call: CallbackQuery):
    emp_id = int(call.data.split(":")[1])
    emp = get_employee_by_id(emp_id)
    pos_id = emp["position_id"] if emp and emp["position_id"] else 0
    all_emps = [e for e in get_all_employees(active_only=True) if e["telegram_id"] > 0]

    if pos_id:
        pos = get_position(pos_id)
        emps = [e for e in all_emps if e["position_id"] == pos_id]
        pos_label = pos["name"] if pos else "—"
    else:
        emps = [e for e in all_emps if not e["position_id"]]
        pos_label = "Lavozim belgilanmagan"

    await call.message.edit_text(
        texts.EMP_DATA_PICK_EMP.format(pos_name=pos_label, count=len(emps)),
        reply_markup=kb.emp_in_position_kb(emps, pos_id)
    )
    await call.answer()


@router.callback_query(F.data.startswith("empdata_excel:"))
async def emp_excel_single(call: CallbackQuery):
    """Tanlangan xodimning oylik Excel hisoboti."""
    if not _is_admin(call.from_user.id):
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    emp_id = int(call.data.split(":")[1])
    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("❌ Xodim topilmadi", show_alert=True)
        return

    await call.answer("⏳ Excel tayyorlanmoqda...")
    now = tz_now()
    year, month = now.year, now.month

    try:
        file_bytes = await asyncio.to_thread(_build_emp_excel, emp_id, year, month)
    except Exception as e:
        logger.error(f"emp_excel_single xato: {e}")
        await call.message.answer("❌ Excel yaratishda xato yuz berdi.")
        return

    fname = f"{emp['full_name'].replace(' ', '_')}_{year}_{month:02d}.xlsx"
    await call.message.answer_document(
        BufferedInputFile(file_bytes, filename=fname),
        caption=f"📊 <b>{emp['full_name']} — {texts.MONTHS_UZ[month]} {year}</b>",
    )


@router.message(F.text == texts.BTN_ADMIN_EMP_EXCEL)
async def emp_excel_all(message: Message):
    """Barcha faol xodimlar — har biri alohida sheet, bitta Excel fayl."""
    if not _is_admin(message.from_user.id):
        return
    await message.answer("⏳ Excel fayl tayyorlanmoqda...")
    now = tz_now()
    year, month = now.year, now.month

    try:
        file_bytes = await asyncio.to_thread(_build_all_emp_excel, year, month)
    except Exception as e:
        logger.error(f"emp_excel_all xato: {e}")
        await message.answer("❌ Excel yaratishda xato yuz berdi.")
        return

    fname = f"xodimlar_ish_haqqi_{year}_{month:02d}.xlsx"
    await message.answer_document(
        BufferedInputFile(file_bytes, filename=fname),
        caption=f"📊 <b>Barcha xodimlar — {texts.MONTHS_UZ[month]} {year}</b>",
    )
