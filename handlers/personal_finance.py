"""Shaxsiy moliya — Boss va Bosh Admin uchun shaxsiy kirim/chiqim hisobkitob."""
import io
import asyncio
import calendar
import logging
from datetime import datetime
from html import escape as _esc

from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, BufferedInputFile

import texts
import keyboards as kb
from database import (
    get_employee_by_telegram_id,
    pf_add_entry, pf_get_monthly,
    pf_get_summary, pf_get_entry, pf_delete_entry, pf_balance_before,
    pf_get_by_date, pf_get_today_totals,
    ensure_owner_pf_categories, get_pf_category_by_ckey, pf_category_label,
)
from states import PersonalFinance
from tzutil import now as tz_now, fmt as fmt_local, last_months, nav_ym
from aiogram.exceptions import TelegramBadRequest

logger = logging.getLogger(__name__)
router = Router()

ARCHIVE_MONTHS = 6


def _can_use(emp) -> bool:
    """Boss, Bosh Admin yoki pf_access berilgan xodim."""
    if not emp:
        return False
    try:
        if emp["role"] in ("boss", "bosh_admin"):
            return True
    except (KeyError, IndexError):
        pass
    try:
        return bool(emp["pf_access"])
    except (KeyError, IndexError):
        return False


def _from_finance(emp) -> bool:
    """PF'ga Moliya bo'limi orqali kirganmi? (Boss/Bosh Admin menyusida PF tugmasi yo'q)"""
    try:
        return emp["role"] in ("boss", "bosh_admin")
    except (KeyError, IndexError):
        return False


def _pf_kb(message: Message):
    """PF menyusi — ortga tugmasi kirish nuqtasiga qarab tanlanadi."""
    me = get_employee_by_telegram_id(message.from_user.id)
    return kb.pf_menu_kb(from_finance=_from_finance(me))


def _parse_ym(raw: str):
    """'2026-07' → (2026, 7), xato bo'lsa None."""
    try:
        y, m = raw.split("-")
        y, m = int(y), int(m)
    except (ValueError, AttributeError):
        return None
    return (y, m) if 1 <= m <= 12 else None


def _parse_date(raw: str):
    """DD.MM.YYYY formatini qabul qiladi, datetime.date qaytaradi yoki None."""
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    return None


# ─── Menyu ─────────────────────────────────────────────────────────────────

@router.message(F.text == texts.BTN_PERSONAL_FINANCE)
async def pf_open(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use(me):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()
    await message.answer(texts.PF_MENU, reply_markup=_pf_kb(message))


@router.message(F.text == texts.BTN_PF_BACK_FINANCE)
async def pf_back_to_finance(message: Message, state: FSMContext):
    """Moliya orqali kirganlar uchun ortga tugmasi. Oddiy xodimda BTN_BACK
    turadi — uni common.py asosiy menyuga qaytaradi."""
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _from_finance(me):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()
    await message.answer(texts.FINANCE_MENU, reply_markup=kb.finance_menu_kb())


# ─── Kirim boshlash ─────────────────────────────────────────────────────────

@router.message(F.text == texts.BTN_PF_INCOME)
async def pf_income_start(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use(me):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()
    await state.update_data(pf_emp_id=me["id"], pf_type="income")
    ensure_owner_pf_categories(me["id"])
    await message.answer(texts.PF_PICK_CAT_INCOME,
                         reply_markup=kb.pf_income_cats_kb(me["id"]))
    await state.set_state(PersonalFinance.choosing_category)


# ─── Chiqim boshlash ────────────────────────────────────────────────────────

@router.message(F.text == texts.BTN_PF_EXPENSE)
async def pf_expense_start(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use(me):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()
    await state.update_data(pf_emp_id=me["id"], pf_type="expense")
    ensure_owner_pf_categories(me["id"])
    await message.answer(texts.PF_PICK_CAT_EXPENSE,
                         reply_markup=kb.pf_expense_cats_kb(me["id"]))
    await state.set_state(PersonalFinance.choosing_category)


# ─── Kategoriya tanlash ─────────────────────────────────────────────────────

@router.callback_query(PersonalFinance.choosing_category, F.data.startswith("pf_cat:"))
async def pf_cat_chosen(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":", 2)
    # "pf_cat:cancel" 2 qismli keladi — uzunlik tekshiruvidan oldin ushlanadi
    if parts[-1] == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return
    if len(parts) < 3:
        await call.answer()
        return
    _, entry_type, cat_key = parts

    data = await state.get_data()
    cat = get_pf_category_by_ckey(data.get("pf_emp_id"), entry_type, cat_key)
    if not cat:
        await call.answer("Noma'lum kategoriya", show_alert=True)
        return

    emoji, cat_name = cat["emoji"], cat["name"]
    await state.update_data(pf_cat_key=cat_key, pf_cat_name=f"{emoji} {cat_name}")

    await call.message.edit_text(
        texts.PF_ENTER_AMOUNT.format(cat=f"{emoji} {cat_name}")
    )
    await state.set_state(PersonalFinance.entering_amount)
    await call.answer()


# ─── Summa kiritish ─────────────────────────────────────────────────────────

@router.message(PersonalFinance.entering_amount, F.text)
async def pf_amount_entered(message: Message, state: FSMContext):
    if message.text in (texts.BTN_CANCEL, texts.BTN_BACK):
        await state.clear()
        await message.answer(texts.CANCELLED, reply_markup=_pf_kb(message))
        return
    raw = message.text.replace(" ", "").replace(",", "").replace(".", "")
    if not raw.isdigit() or int(raw) <= 0:
        await message.answer(texts.PF_AMOUNT_INVALID)
        return
    await state.update_data(pf_amount=int(raw))
    await message.answer(texts.PF_ENTER_NOTE, reply_markup=kb.pf_note_kb())
    await state.set_state(PersonalFinance.entering_note)


# ─── Izoh kiritish ──────────────────────────────────────────────────────────

@router.message(PersonalFinance.entering_note, F.text)
async def pf_note_entered(message: Message, state: FSMContext):
    if message.text in (texts.BTN_CANCEL, texts.BTN_BACK):
        await state.clear()
        await message.answer(texts.CANCELLED, reply_markup=_pf_kb(message))
        return
    note = "" if message.text == texts.BTN_PF_NOTE_SKIP else message.text
    await state.update_data(pf_note=note)
    await message.answer(texts.PF_ENTER_DATE, reply_markup=kb.pf_date_kb())
    await state.set_state(PersonalFinance.entering_date)


# ─── Sana kiritish ──────────────────────────────────────────────────────────

@router.message(PersonalFinance.entering_date, F.text)
async def pf_date_entered(message: Message, state: FSMContext):
    raw = message.text.strip()
    if raw in (texts.BTN_CANCEL, texts.BTN_BACK):
        await state.clear()
        await message.answer(texts.CANCELLED, reply_markup=_pf_kb(message))
        return
    if raw == texts.BTN_PF_TODAY:
        entry_date = tz_now().date()
    else:
        entry_date = _parse_date(raw)
        if not entry_date:
            await message.answer(texts.PF_DATE_INVALID)
            return

    data = await state.get_data()
    await state.clear()

    emp_id = data["pf_emp_id"]
    entry_type = data["pf_type"]
    cat_key = data["pf_cat_key"]
    cat_name = data["pf_cat_name"]
    amount = data["pf_amount"]
    note = data.get("pf_note", "")

    pf_add_entry(emp_id, entry_type, cat_key, amount, note, str(entry_date))
    sign = "+" if entry_type == "income" else "-"
    await message.answer(
        texts.PF_SAVED.format(sign=sign, amount=amount, cat=cat_name),
        reply_markup=_pf_kb(message)
    )


# ─── Oy hisoboti (joriy oy va arxiv uchun umumiy) ───────────────────────────

def _pf_summary_text(emp_id: int, year: int, month: int, with_today: bool):
    """Oy xulosasi matni. Yozuv bo'lmasa None.

    with_today — bugungi blok va kunlik byudjet (faqat joriy oy uchun ma'noli).
    """
    summary = pf_get_summary(emp_id, year, month)
    if not summary["income"] and not summary["expense"]:
        return None

    out = texts.PF_SUMMARY_HEADER.format(month=texts.MONTHS_UZ[month], year=year)

    # Kirim
    if summary["income"]:
        out += texts.PF_SUMMARY_INCOME.format(total=summary["income"])
        for (etype, cat_key), amount in summary["by_cat"].items():
            if etype != "income":
                continue
            emoji, name = (pf_category_label(cat_key)
                           or texts.PF_INCOME_CATS.get(cat_key, ("📋", cat_key)))
            out += texts.PF_SUMMARY_LINE.format(emoji=emoji, name=name, amount=amount)

    # Chiqim
    if summary["expense"]:
        out += texts.PF_SUMMARY_EXPENSE.format(total=summary["expense"])
        for (etype, cat_key), amount in summary["by_cat"].items():
            if etype != "expense":
                continue
            emoji, name = (pf_category_label(cat_key)
                           or texts.PF_EXPENSE_CATS.get(cat_key, ("📋", cat_key)))
            out += texts.PF_SUMMARY_LINE.format(emoji=emoji, name=name, amount=amount)

    # Sof
    net = summary["net"]
    if net >= 0:
        out += texts.PF_SUMMARY_NET_PLUS.format(net=net)
    else:
        out += texts.PF_SUMMARY_NET_MINUS.format(net=abs(net))

    if not with_today:
        return out

    # Bugungi chiqim + kunlik byudjet — faqat joriy oy uchun
    now = tz_now()
    today = pf_get_today_totals(emp_id, now.strftime("%Y-%m-%d"))
    if today["expense"] > 0:
        out += texts.PF_SUMMARY_TODAY.format(today=today["expense"])
    else:
        out += texts.PF_SUMMARY_TODAY_NONE

    remaining_days = max(calendar.monthrange(year, month)[1] - now.day, 1)
    if net <= 0:
        out += texts.PF_SUMMARY_NO_LIMIT
    else:
        out += texts.PF_SUMMARY_BUDGET.format(
            qoldiq=net, days=remaining_days,
            limit=net // remaining_days
        )
    return out


def _pfsum_view(me, year: int, month: int):
    """PF oy hisoboti: matn + oy navigatsiyasi."""
    now = tz_now()
    is_cur = (year, month) == (now.year, now.month)
    out = _pf_summary_text(me["id"], year, month, with_today=is_cur)
    return out or texts.PF_SUMMARY_EMPTY, kb.month_nav_kb("pfsum", year, month)


@router.message(F.text == texts.BTN_PF_SUMMARY)
async def pf_summary(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use(me):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()

    now = tz_now()
    text, markup = _pfsum_view(me, now.year, now.month)
    await message.answer(text, reply_markup=markup)


@router.callback_query(F.data.startswith("pfsum:"))
async def pf_summary_nav(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not _can_use(me):
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    year, month = nav_ym(call.data)
    text, markup = _pfsum_view(me, year, month)
    try:
        await call.message.edit_text(text, reply_markup=markup)
    except TelegramBadRequest:
        pass  # matn o'zgarmagan bo'lsa e'tiborsiz
    await call.answer()


# ─── Excel hisobot (joriy oy va arxiv uchun umumiy) ─────────────────────────

async def _send_pf_excel(message: Message, emp, year: int, month: int) -> bool:
    """Tanlangan oy uchun PF Excel yuboradi. Yozuv bo'lmasa False."""
    entries = pf_get_monthly(emp["id"], year, month)
    if not entries:
        return False

    opening = pf_balance_before(emp["id"], year, month)
    file_bytes = await asyncio.to_thread(
        _build_pf_excel, entries, year, month, emp["full_name"], opening
    )
    await message.answer_document(
        BufferedInputFile(
            file_bytes, filename=f"shaxsiy_moliya_{year}_{month:02d}.xlsx"
        ),
        caption=f"📥 Shaxsiy moliya — {texts.MONTHS_UZ[month]} {year}"
    )
    return True


@router.message(F.text == texts.BTN_PF_EXCEL)
async def pf_excel(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use(me):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()

    now = tz_now()
    await message.answer(
        texts.MONTH_EXCEL_PROMPT.format(
            month=texts.MONTHS_UZ[now.month], year=now.year
        ),
        reply_markup=kb.month_excel_kb("pfxl", now.year, now.month)
    )


@router.callback_query(F.data.startswith("pfxl:"))
async def pf_excel_nav(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not _can_use(me):
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    year, month = nav_ym(call.data)
    try:
        await call.message.edit_text(
            texts.MONTH_EXCEL_PROMPT.format(
                month=texts.MONTHS_UZ[month], year=year
            ),
            reply_markup=kb.month_excel_kb("pfxl", year, month)
        )
    except TelegramBadRequest:
        pass
    await call.answer()


@router.callback_query(F.data.startswith("pfxldl:"))
async def pf_excel_download(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not _can_use(me):
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return
    year, month = nav_ym(call.data)
    await call.answer("⏳ Tayyorlanmoqda...")
    if not await _send_pf_excel(call.message, me, year, month):
        await call.message.answer(texts.ARCHIVE_MONTH_EXCEL_EMPTY.format(
            month=texts.MONTHS_UZ[month], year=year
        ))


# ─── Arxiv (o'tgan oylar — faqat o'qish) ────────────────────────────────────

@router.message(F.text == texts.BTN_PF_ARCHIVE)
async def pf_archive(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use(me):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()
    await message.answer(
        texts.ARCHIVE_PICK_MONTH,
        reply_markup=kb.archive_months_kb("pf_arc", last_months(ARCHIVE_MONTHS))
    )


@router.callback_query(F.data.startswith("pf_arc:"))
async def pf_archive_month(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not _can_use(me):
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    raw = call.data.split(":", 1)[1]
    if raw == "cancel":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return
    ym = _parse_ym(raw)
    if not ym:
        await call.answer("❌ Xato", show_alert=True)
        return
    year, month = ym

    out = _pf_summary_text(me["id"], year, month, with_today=False)
    await call.message.edit_text(
        out or texts.PF_SUMMARY_EMPTY,
        reply_markup=kb.archive_excel_kb("pf_arcx", year, month)
    )
    await call.answer()


@router.callback_query(F.data.startswith("pf_arcx:"))
async def pf_archive_excel(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not _can_use(me):
        await call.answer(texts.NO_PERMISSION, show_alert=True)
        return

    ym = _parse_ym(call.data.split(":", 1)[1])
    if not ym:
        await call.answer("❌ Xato", show_alert=True)
        return
    year, month = ym

    await call.answer("⏳ Tayyorlanmoqda...")
    if not await _send_pf_excel(call.message, me, year, month):
        await call.message.answer(texts.ARCHIVE_MONTH_EXCEL_EMPTY.format(
            month=texts.MONTHS_UZ[month], year=year
        ))


def _build_pf_excel(entries, year: int, month: int,
                    owner_name: str, opening: int = 0) -> bytes:
    """Shaxsiy moliya Excel — biznes moliya (finance) formati bilan bir xil."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Qoldiq to'g'ri yig'ilishi uchun ASC (sana, id)
    entries = sorted(entries, key=lambda e: (str(e["entry_date"]), e["id"]))

    def _cat(cat):
        info = (pf_category_label(cat)
                or texts.PF_ALL_CATS.get(cat, ("📋", cat)))
        return f"{info[0]} {info[1]}"

    def _sana(e):
        try:
            return datetime.strptime(str(e["entry_date"])[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            return str(e["entry_date"])

    def _vaqt(e):
        try:
            return fmt_local(e["created_at"], "%H:%M")
        except Exception:
            return ""

    wb = Workbook()
    ws = wb.active
    ws.title = "Yozuvlar"

    title = ws.cell(row=1, column=1,
                    value=f"📊 Shaxsiy — {owner_name} · {texts.MONTHS_UZ[month]} {year}")
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 22

    headers = ["Sana", "Vaqt", "Kategoriya", "Izoh",
               "Rasxod (so'm)", "Kirim (so'm)", "Qoldiq (so'm)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # Boshlang'ich qoldiq (oldingi oylardan)
    pm = month - 1 if month > 1 else 12
    open_row = 4
    open_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    ws.cell(row=open_row, column=3,
            value=f"📦 {texts.MONTHS_UZ[pm]} oyidan qolgan qoldiq")
    oc = ws.cell(row=open_row, column=7, value=opening)
    oc.number_format = '#,##0'
    oc.font = Font(bold=True)
    for c in range(1, 8):
        cell = ws.cell(row=open_row, column=c)
        cell.border = border
        cell.fill = open_fill

    # Yozuvlar — har qatorda tirik Qoldiq formulasi
    start = open_row + 1
    for idx, e in enumerate(entries):
        i = start + idx
        ws.cell(row=i, column=1, value=_sana(e))
        ws.cell(row=i, column=2, value=_vaqt(e))
        ws.cell(row=i, column=3, value=_cat(e["category"]))
        ws.cell(row=i, column=4, value=e["note"] or "")
        if e["entry_type"] == "expense":
            ac = ws.cell(row=i, column=5, value=e["amount"])
            ac.number_format = '#,##0'
            ac.font = Font(color="8B0000")
        else:
            ac = ws.cell(row=i, column=6, value=e["amount"])
            ac.number_format = '#,##0'
            ac.font = Font(color="006400")
        qc = ws.cell(row=i, column=7, value=f"=G{i - 1}+F{i}-E{i}")
        qc.number_format = '#,##0'
        qc.font = Font(bold=True)
        for c in range(1, 8):
            ws.cell(row=i, column=c).border = border

    # JAMI qatori
    last = start + len(entries) - 1
    trow = last + 1
    ws.cell(row=trow, column=4, value="JAMI").font = Font(bold=True)
    te = ws.cell(row=trow, column=5, value=f"=SUM(E{start}:E{last})")
    te.number_format = '#,##0'
    te.font = Font(bold=True, color="8B0000")
    ti = ws.cell(row=trow, column=6, value=f"=SUM(F{start}:F{last})")
    ti.number_format = '#,##0'
    ti.font = Font(bold=True, color="006400")
    tq = ws.cell(row=trow, column=7, value=f"=G{open_row}+F{trow}-E{trow}")
    tq.number_format = '#,##0'
    tq.font = Font(bold=True)
    for c in range(1, 8):
        ws.cell(row=trow, column=c).border = border
        ws.cell(row=trow, column=c).fill = open_fill

    for col, w in zip("ABCDEFG", [12, 8, 28, 36, 15, 15, 16]):
        ws.column_dimensions[col].width = w

    # Sheet 2: Xulosa (turkumlar bo'yicha)
    ws2 = wb.create_sheet("Xulosa")
    for col, h in enumerate(["Tur", "Turkum", "Yozuvlar", "Jami (so'm)"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = Font(bold=True, color="FFFFFF")
        c.border = border

    agg = {}
    inc_total = exp_total = 0
    for e in entries:
        k = (e["entry_type"], e["category"])
        a = agg.setdefault(k, [0, 0])
        a[0] += 1
        a[1] += e["amount"]
        if e["entry_type"] == "income":
            inc_total += e["amount"]
        else:
            exp_total += e["amount"]

    row = 2
    for tp in ("income", "expense"):
        type_label = "Kirim" if tp == "income" else "Chiqim"
        for (t, cat), (cnt, total) in agg.items():
            if t != tp:
                continue
            ws2.cell(row=row, column=1, value=type_label)
            ws2.cell(row=row, column=2, value=_cat(cat))
            ws2.cell(row=row, column=3, value=cnt)
            ac = ws2.cell(row=row, column=4, value=total)
            ac.number_format = '#,##0'
            ac.font = Font(color="006400" if tp == "income" else "8B0000")
            for c in range(1, 5):
                ws2.cell(row=row, column=c).border = border
            row += 1

    row += 1
    inc_row = row
    ws2.cell(row=row, column=1, value="Jami kirim").font = Font(bold=True)
    c = ws2.cell(row=row, column=4, value=inc_total)
    c.number_format = '#,##0'; c.font = Font(bold=True, color="006400")
    row += 1
    exp_row = row
    ws2.cell(row=row, column=1, value="Jami chiqim").font = Font(bold=True)
    c = ws2.cell(row=row, column=4, value=exp_total)
    c.number_format = '#,##0'; c.font = Font(bold=True, color="8B0000")
    row += 1
    net = inc_total - exp_total
    ws2.cell(row=row, column=1, value="Sof natija").font = Font(bold=True, size=12)
    c = ws2.cell(row=row, column=4, value=f"=D{inc_row}-D{exp_row}")
    c.number_format = '#,##0'
    c.font = Font(bold=True, color="006400" if net >= 0 else "8B0000")

    for col, w in zip("ABCD", [22, 26, 12, 18]):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Yozuvni o'chirish ──────────────────────────────────────────────────────

def _pf_day_list(entries) -> str:
    """Bir kundagi yozuvlar — to'liq tafsilotli matn (vaqt, kategoriya, summa, izoh)."""
    lines = []
    for i, e in enumerate(entries, 1):
        emoji, name = (pf_category_label(e["category"])
                       or texts.PF_ALL_CATS.get(e["category"], ("📋", e["category"])))
        sign = "➕" if e["entry_type"] == "income" else "➖"
        try:
            t = fmt_local(e["created_at"], "%H:%M")
        except Exception:
            t = ""
        note = f" — {_esc(e['note'])}" if e["note"] else ""
        lines.append(
            f"{i}. {sign} <b>{e['amount']:,}</b> · {emoji}{name}"
            + (f" · {t}" if t else "") + note
        )
    return "\n".join(lines)


@router.message(F.text == texts.BTN_PF_DELETE)
async def pf_delete_start(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use(me):
        await message.answer(texts.NO_PERMISSION)
        return
    await state.clear()
    await state.update_data(pf_emp_id=me["id"])
    await message.answer(texts.PF_DELETE_ASK_DATE, reply_markup=kb.pf_date_kb())
    await state.set_state(PersonalFinance.deleting_date)


@router.message(PersonalFinance.deleting_date, F.text)
async def pf_delete_date(message: Message, state: FSMContext):
    if message.text in (texts.BTN_CANCEL, texts.BTN_BACK):
        await state.clear()
        await message.answer(texts.CANCELLED, reply_markup=_pf_kb(message))
        return

    raw = message.text.strip()
    if raw == texts.BTN_PF_TODAY:
        d = tz_now().date()
    else:
        d = _parse_date(raw)
        if not d:
            await message.answer(texts.PF_DATE_INVALID, reply_markup=kb.pf_date_kb())
            return

    data = await state.get_data()
    emp_id = data.get("pf_emp_id")
    date_str = d.strftime("%Y-%m-%d")
    label = d.strftime("%d.%m.%Y")
    entries = pf_get_by_date(emp_id, date_str)
    if not entries:
        await message.answer(texts.PF_DELETE_DAY_EMPTY.format(date=label),
                             reply_markup=kb.pf_date_kb())
        return

    await state.update_data(pf_del_date=date_str)
    await state.set_state(PersonalFinance.deleting)
    await message.answer(
        texts.PF_DELETE_PICK_DAY.format(date=label, list=_pf_day_list(entries)),
        reply_markup=kb.pf_entries_kb(entries)
    )
    await message.answer("👇 Yuqoridan tanlang yoki Bekor qiling.",
                         reply_markup=_pf_kb(message))


@router.callback_query(PersonalFinance.deleting, F.data.startswith("pf_del:"))
async def pf_delete_confirm(call: CallbackQuery, state: FSMContext):
    raw = call.data.split(":", 1)[1]
    if raw == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    try:
        entry_id = int(raw)
    except ValueError:
        await call.answer("Xato", show_alert=True)
        return

    data = await state.get_data()
    emp_id = data.get("pf_emp_id")
    date_str = data.get("pf_del_date")
    row = pf_get_entry(entry_id, emp_id)
    if not row:
        await call.answer("Topilmadi yoki ruxsat yo'q.", show_alert=True)
        return

    pf_delete_entry(entry_id, emp_id)
    await call.answer("✅ O'chirildi")

    # Shu kunda qolgan yozuvlarni qayta ko'rsatamiz — ketma-ket o'chirish uchun
    remaining = pf_get_by_date(emp_id, date_str) if date_str else []
    if remaining:
        try:
            label = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            label = str(date_str)
        await call.message.edit_text(
            texts.PF_DELETE_PICK_DAY.format(date=label, list=_pf_day_list(remaining)),
            reply_markup=kb.pf_entries_kb(remaining)
        )
    else:
        await state.clear()
        await call.message.edit_text(texts.PF_DELETED)
