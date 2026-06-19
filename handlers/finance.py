"""Moliya bo'limi — Bosh Admin va Boss uchun.

Yangi oqim:
  Kirim: Podachot | Sotuvdan tushum | Boshqa → Summa → Izoh → Saqlash
  Chiqim: Ovqat | Yo'lkira | Ta'minot | Xarajat | Boshqa → Summa → Izoh → Saqlash
           Hodimlar uchun Avans → Xodim tanlash → Summa → Izoh → Saqlash
"""
import io
import logging
import asyncio
from datetime import datetime, timedelta

from aiogram import Router, F, Bot
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, BufferedInputFile

import texts
import keyboards as kb
from states import FinanceEntry, FinanceDelete
from database import (
    get_employee_by_telegram_id,
    get_all_employees,
    get_employee_by_id,
    create_finance_entry,
    get_monthly_finance_entries,
    get_monthly_finance_summary,
    add_salary_entry,
    is_month_closed,
    get_finance_entry,
    get_finance_entries_by_date,
    delete_finance_entry,
    get_finance_balance,
    get_finance_balance_before,
)
from tzutil import now as tz_now, fmt as fmt_local

logger = logging.getLogger(__name__)
router = Router()


def _can_use_finance(emp) -> bool:
    if not emp:
        return False
    try:
        return emp["role"] in ("boss", "bosh_admin")
    except (KeyError, IndexError):
        return False


def _back_kb(emp):
    try:
        role = emp["role"]
    except (KeyError, IndexError):
        role = "employee"
    if role == "bosh_admin":
        return kb.main_menu_kb(is_bosh_admin=True)
    if role == "boss":
        return kb.main_menu_kb(is_boss=True)
    return kb.main_menu_kb(is_admin=bool(emp["is_admin"]))


# ===== Kirish =====

@router.message(F.text == texts.BTN_BOSS_FINANCE)
async def finance_open(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use_finance(me):
        await message.answer(texts.FINANCE_NO_PERMISSION)
        return
    await state.clear()
    await message.answer(texts.FINANCE_MENU, reply_markup=kb.finance_menu_kb())


# ===== Kirim / Chiqim boshlash =====

@router.message(F.text == texts.BTN_FINANCE_INCOME)
async def finance_add_income_start(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use_finance(me):
        await message.answer(texts.FINANCE_NO_PERMISSION)
        return
    await state.clear()
    await message.answer(
        texts.FINANCE_PICK_CATEGORY_INCOME,
        reply_markup=kb.finance_categories_kb("income")
    )


@router.message(F.text == texts.BTN_FINANCE_EXPENSE)
async def finance_add_expense_start(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use_finance(me):
        await message.answer(texts.FINANCE_NO_PERMISSION)
        return
    await state.clear()
    await message.answer(
        texts.FINANCE_PICK_CATEGORY_EXPENSE,
        reply_markup=kb.finance_categories_kb("expense")
    )


# ===== Kategoriya tanlash =====

@router.callback_query(F.data.startswith("fin_cat:"))
async def finance_category_chosen(call: CallbackQuery, state: FSMContext):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not _can_use_finance(me):
        await call.answer(texts.FINANCE_NO_PERMISSION, show_alert=True)
        return

    parts = call.data.split(":")
    if len(parts) != 3:
        await call.answer("Xato", show_alert=True)
        return
    _, entry_type, cat_key = parts

    if cat_key == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return

    # Shaxsiy xarajatlar — ichki turkumlar menyusi
    if cat_key == "personal" and entry_type == "expense":
        await call.message.edit_text(
            texts.FINANCE_PICK_PERSONAL,
            reply_markup=kb.finance_personal_cats_kb()
        )
        await call.answer()
        return

    # Shaxsiy menyudan ortga — chiqim turkumlariga qaytish
    if cat_key == "backexp":
        await call.message.edit_text(
            texts.FINANCE_PICK_CATEGORY_EXPENSE,
            reply_markup=kb.finance_categories_kb("expense")
        )
        await call.answer()
        return

    # Kategoriya lug'atidan olish
    all_cats = texts.FINANCE_CATEGORIES
    if cat_key not in all_cats:
        await call.answer("Noma'lum turkum", show_alert=True)
        return

    emoji, cat_name = all_cats[cat_key]
    type_name = "Kirim" if entry_type == "income" else "Chiqim"

    await state.update_data(
        fin_owner=me["id"],
        fin_type=entry_type,
        fin_cat_key=cat_key,
        fin_cat_name=cat_name,
        fin_cat_emoji=emoji,
        fin_linked_emp_id=None,
        fin_linked_emp_name=None,
    )
    await call.answer()

    # Avans bo'lsa — xodim tanlash
    if cat_key == "advance" and entry_type == "expense":
        now = tz_now()
        if is_month_closed(now.year, now.month):
            await call.message.edit_text(
                texts.FINANCE_ADVANCE_MONTH_CLOSED.format(
                    month=texts.MONTHS_UZ[now.month], year=now.year
                )
            )
            return
        employees = get_all_employees(active_only=True)
        # Boss va Bosh Admin hisoblanmaydi
        employees = [e for e in employees if e["role"] not in ("boss", "bosh_admin")]
        if not employees:
            await call.message.edit_text(
                "❌ Aktiv xodimlar yo'q.",
            )
            return
        await call.message.edit_text(
            texts.FINANCE_PICK_EMPLOYEE_ADVANCE,
            reply_markup=kb.finance_employees_kb(employees)
        )
        await state.set_state(FinanceEntry.selecting_employee)
        return

    # Oddiy kategoriya — summaga o'tish
    await call.message.edit_text(
        texts.FINANCE_ASK_AMOUNT.format(
            emoji="➕" if entry_type == "income" else "➖",
            type_name=type_name,
            category=f"{emoji} {cat_name}",
        )
    )
    await state.set_state(FinanceEntry.entering_amount)


# ===== Avans: xodim tanlash =====

@router.callback_query(FinanceEntry.selecting_employee, F.data.startswith("fin_emp:"))
async def finance_employee_chosen(call: CallbackQuery, state: FSMContext):
    emp_id_str = call.data.split(":")[-1]

    if emp_id_str == "cancel":
        await state.clear()
        await call.message.edit_text(texts.CANCELLED)
        await call.message.answer(texts.FINANCE_MENU, reply_markup=kb.finance_menu_kb())
        await call.answer()
        return

    try:
        emp_id = int(emp_id_str)
    except ValueError:
        await call.answer("Xato ID", show_alert=True)
        return

    emp = get_employee_by_id(emp_id)
    if not emp:
        await call.answer("Xodim topilmadi", show_alert=True)
        return

    data = await state.get_data()
    await state.update_data(
        fin_linked_emp_id=emp_id,
        fin_linked_emp_name=emp["full_name"],
    )
    await call.answer()

    type_name = "Chiqim"
    cat_name = data["fin_cat_name"]
    emoji = data["fin_cat_emoji"]

    await call.message.edit_text(
        f"👤 Xodim: <b>{emp['full_name']}</b>\n\n"
        + texts.FINANCE_ASK_AMOUNT.format(
            emoji="➖",
            type_name=type_name,
            category=f"{emoji} {cat_name}",
        )
    )
    await state.set_state(FinanceEntry.entering_amount)


# ===== Summa kiritish =====

@router.message(FinanceEntry.entering_amount, F.text)
async def finance_amount_handler(message: Message, state: FSMContext):
    if message.text == texts.BTN_CANCEL:
        await state.clear()
        await message.answer(texts.CANCELLED, reply_markup=kb.finance_menu_kb())
        return
    try:
        cleaned = message.text.replace(" ", "").replace(",", "").replace(".", "").strip()
        amount = int(cleaned)
        if amount <= 0 or amount > 10_000_000_000:
            raise ValueError()
    except ValueError:
        await message.answer(texts.FINANCE_AMOUNT_INVALID)
        return

    data = await state.get_data()
    type_name = "Kirim" if data["fin_type"] == "income" else "Chiqim"
    await state.update_data(fin_amount=amount)
    await message.answer(
        texts.FINANCE_ASK_NOTE.format(
            emoji="➕" if data["fin_type"] == "income" else "➖",
            type_name=type_name,
            category=f"{data['fin_cat_emoji']} {data['fin_cat_name']}",
            amount=amount,
        ),
        reply_markup=kb.finance_note_skip_kb()
    )
    await state.set_state(FinanceEntry.entering_note)


# ===== Izoh → Sana → Saqlash =====

@router.message(FinanceEntry.entering_note, F.text)
async def finance_note_handler(message: Message, state: FSMContext):
    if message.text in (texts.BTN_CANCEL, texts.BTN_BACK):
        await state.clear()
        await message.answer(texts.CANCELLED, reply_markup=kb.finance_menu_kb())
        return
    note = None
    if message.text != texts.BTN_FINANCE_NOTE_SKIP:
        stripped = message.text.strip()
        if stripped:
            note = stripped[:500] if len(stripped) > 500 else stripped
    await state.update_data(fin_note=note)
    await message.answer(texts.FINANCE_ASK_DATE, reply_markup=kb.finance_date_kb())
    await state.set_state(FinanceEntry.entering_date)


def _parse_user_date(text: str):
    """KK.OO.YYYY → datetime.date yoki None."""
    cleaned = text.strip().replace("/", ".").replace("-", ".")
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    return None


@router.message(FinanceEntry.entering_date, F.text)
async def finance_date_handler(message: Message, state: FSMContext, bot: Bot):
    from html import escape as _esc
    if message.text in (texts.BTN_CANCEL, texts.BTN_BACK):
        await state.clear()
        await message.answer(texts.CANCELLED, reply_markup=kb.finance_menu_kb())
        return

    entry_date_utc = None  # None — hozirgi vaqt
    if message.text == texts.BTN_FINANCE_TODAY:
        chosen = tz_now()
        when_str = chosen.strftime("%d.%m.%Y %H:%M")
    else:
        d = _parse_user_date(message.text)
        if not d:
            await message.answer(texts.FINANCE_DATE_INVALID,
                                 reply_markup=kb.finance_date_kb())
            return
        # Mahalliy (Toshkent) 12:00 → UTC = −5 soat
        chosen = datetime(d.year, d.month, d.day, 12, 0, 0)
        entry_date_utc = (chosen - timedelta(hours=5)).strftime("%Y-%m-%d %H:%M:%S")
        when_str = chosen.strftime("%d.%m.%Y")

    data = await state.get_data()

    # Avans: tanlangan oy yopiq bo'lsa — bloklash
    if data.get("fin_cat_key") == "advance" and data.get("fin_linked_emp_id"):
        if is_month_closed(chosen.year, chosen.month):
            await state.clear()
            await message.answer(
                texts.FINANCE_ADVANCE_MONTH_CLOSED.format(
                    month=texts.MONTHS_UZ[chosen.month], year=chosen.year
                ),
                reply_markup=kb.finance_menu_kb()
            )
            return

    note = data.get("fin_note")
    await state.clear()
    try:
        entry_id = create_finance_entry(
            owner_id=data["fin_owner"],
            entry_type=data["fin_type"],
            category=data["fin_cat_key"],
            amount=int(data["fin_amount"]),
            note=note,
            linked_employee_id=data.get("fin_linked_emp_id"),
            entry_date_utc=entry_date_utc,
        )
        logger.info("Finance entry %s yaratildi: owner=%s, %s %s",
                    entry_id, data["fin_owner"], data["fin_type"], data["fin_amount"])

        # Avans: xodimning ish haqqidan ham chegirish + bildirishnoma
        salary_note_line = ""
        if data["fin_cat_key"] == "advance" and data.get("fin_linked_emp_id"):
            reason = note or "Moliya bo'limidan avans"
            try:
                add_salary_entry(
                    employee_id=data["fin_linked_emp_id"],
                    entry_type="avans",
                    amount=int(data["fin_amount"]),
                    reason=reason,
                    created_by=data["fin_owner"],
                )
                salary_note_line = texts.FINANCE_ADVANCE_SALARY_NOTED.format(
                    amount=int(data["fin_amount"])
                )
                emp = get_employee_by_id(data["fin_linked_emp_id"])
                if emp and emp["telegram_id"] > 0:
                    try:
                        await bot.send_message(
                            emp["telegram_id"],
                            texts.NOTIFY_SALARY_ADDED.format(
                                emoji="💸", type_name="Avans", sign="−",
                                amount=int(data["fin_amount"]),
                                reason=_esc(reason),
                            )
                        )
                    except Exception as notify_exc:
                        logger.warning("Avans bildirishnoma xato: tg=%s %s",
                                       emp["telegram_id"], notify_exc)
            except Exception as sal_exc:
                logger.exception("Avans ish haqqi yozuvi xato: %s", sal_exc)

        type_emoji = "➕" if data["fin_type"] == "income" else "➖"
        type_name = "Kirim" if data["fin_type"] == "income" else "Chiqim"
        note_line = texts.FINANCE_NOTE_FRAGMENT.format(note=_esc(note)) if note else ""
        await message.answer(
            texts.FINANCE_SAVED.format(
                type_emoji=type_emoji,
                type_name=type_name,
                cat_emoji=data["fin_cat_emoji"],
                category=data["fin_cat_name"],
                amount=int(data["fin_amount"]),
                when=when_str,
                note_line=note_line,
            ) + salary_note_line,
            reply_markup=kb.finance_menu_kb()
        )
    except KeyError as exc:
        logger.exception("Finance date: state malumotlari yoq -- %s", exc)
        await message.answer(
            "Texnik xato: holat malumotlari yoqoldi. Moliya bolimiga qayta kiring.",
            reply_markup=kb.finance_menu_kb()
        )
    except Exception as exc:
        logger.exception("Finance saqlash xato: %s", exc)
        await message.answer(
            f"Saqlashda xato yuz berdi.\n\n"
            f"<code>{type(exc).__name__}: {_esc(str(exc))}</code>\n\n"
            f"Qayta urinib koring.",
            reply_markup=kb.finance_menu_kb()
        )


# ===== Yozuvni o'chirish (sana orqali) =====

@router.message(F.text == texts.BTN_FINANCE_DELETE)
async def finance_delete_start(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use_finance(me):
        await message.answer(texts.FINANCE_NO_PERMISSION)
        return
    await state.clear()
    await message.answer(texts.FINANCE_DELETE_ASK_DATE,
                         reply_markup=kb.finance_date_kb())
    await state.set_state(FinanceDelete.entering_date)


@router.message(FinanceDelete.entering_date, F.text)
async def finance_delete_date_handler(message: Message, state: FSMContext):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use_finance(me):
        await state.clear()
        await message.answer(texts.FINANCE_NO_PERMISSION)
        return
    if message.text in (texts.BTN_CANCEL, texts.BTN_BACK):
        await state.clear()
        await message.answer(texts.CANCELLED, reply_markup=kb.finance_menu_kb())
        return

    if message.text == texts.BTN_FINANCE_TODAY:
        d = tz_now().date()
    else:
        d = _parse_user_date(message.text)
        if not d:
            await message.answer(texts.FINANCE_DATE_INVALID,
                                 reply_markup=kb.finance_date_kb())
            return

    date_str = d.strftime("%Y-%m-%d")
    label = d.strftime("%d.%m.%Y")
    entries = get_finance_entries_by_date(me["id"], date_str)
    if not entries:
        await message.answer(texts.FINANCE_DELETE_EMPTY.format(date=label),
                             reply_markup=kb.finance_date_kb())
        return

    await state.clear()
    await message.answer(
        texts.FINANCE_DELETE_PICK.format(date=label),
        reply_markup=kb.finance_del_entries_kb(entries)
    )
    await message.answer("👇 Yuqoridan yozuvni tanlang yoki Bekor qiling.",
                         reply_markup=kb.finance_menu_kb())


@router.callback_query(F.data.startswith("fin_del:"))
async def finance_delete_pick(call: CallbackQuery):
    from html import escape as _esc
    me = get_employee_by_telegram_id(call.from_user.id)
    if not _can_use_finance(me):
        await call.answer(texts.FINANCE_NO_PERMISSION, show_alert=True)
        return
    arg = call.data.split(":", 1)[1]
    if arg == "cancel":
        await call.message.edit_text(texts.CANCELLED)
        await call.answer()
        return
    entry = get_finance_entry(int(arg), me["id"])
    if not entry:
        await call.answer("Yozuv topilmadi", show_alert=True)
        return
    cat_info = texts.FINANCE_CATEGORIES.get(entry["category"], ("📋", entry["category"]))
    note_line = (texts.FINANCE_NOTE_FRAGMENT.format(note=_esc(entry["note"]))
                 if entry["note"] else "")
    advance_warn = (texts.FINANCE_DELETE_ADVANCE_WARN
                    if entry["category"] == "advance" else "")
    await call.message.edit_text(
        texts.FINANCE_DELETE_CONFIRM.format(
            type_emoji="➕" if entry["entry_type"] == "income" else "➖",
            type_name="Kirim" if entry["entry_type"] == "income" else "Chiqim",
            cat_emoji=cat_info[0],
            category=cat_info[1],
            amount=entry["amount"],
            when=fmt_local(entry["entry_date"], "%d.%m.%Y %H:%M"),
            note_line=note_line,
            advance_warn=advance_warn,
        ),
        reply_markup=kb.finance_del_confirm_kb(entry["id"])
    )
    await call.answer()


@router.callback_query(F.data.startswith("fin_delc:"))
async def finance_delete_confirm(call: CallbackQuery):
    me = get_employee_by_telegram_id(call.from_user.id)
    if not _can_use_finance(me):
        await call.answer(texts.FINANCE_NO_PERMISSION, show_alert=True)
        return
    entry_id = int(call.data.split(":", 1)[1])
    entry = get_finance_entry(entry_id, me["id"])
    if not entry:
        await call.answer("Yozuv topilmadi", show_alert=True)
        return
    ok = delete_finance_entry(entry_id, me["id"])
    if not ok:
        await call.answer("O'chirib bo'lmadi", show_alert=True)
        return
    cat_info = texts.FINANCE_CATEGORIES.get(entry["category"], ("📋", entry["category"]))
    logger.info("Finance entry %s o'chirildi (owner=%s)", entry_id, me["id"])
    await call.message.edit_text(
        texts.FINANCE_DELETED.format(
            type_emoji="➕" if entry["entry_type"] == "income" else "➖",
            amount=entry["amount"],
            category=cat_info[1],
        )
    )
    await call.answer("O'chirildi")


@router.message(F.text == texts.BTN_FINANCE_SUMMARY)
async def finance_summary(message: Message):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use_finance(me):
        await message.answer(texts.FINANCE_NO_PERMISSION)
        return

    now = tz_now()
    summary = get_monthly_finance_summary(me["id"], now.year, now.month)
    out = texts.FINANCE_SUMMARY_HEADER.format(
        month=texts.MONTHS_UZ[now.month], year=now.year
    )

    if not summary["by_category"]["income"] and not summary["by_category"]["expense"]:
        out += texts.FINANCE_SUMMARY_EMPTY
        out += texts.FINANCE_SUMMARY_BALANCE.format(balance=get_finance_balance(me["id"]))
        await message.answer(out)
        return

    if summary["by_category"]["income"]:
        out += texts.FINANCE_SUMMARY_INCOME.format(total=summary["income_total"])
        for row in summary["by_category"]["income"]:
            cat_info = texts.FINANCE_CATEGORIES.get(row["category"], ("📋", row["category"]))
            out += texts.FINANCE_SUMMARY_CAT_LINE.format(
                emoji=cat_info[0], category=cat_info[1],
                total=row["total"], cnt=row["cnt"]
            )

    if summary["by_category"]["expense"]:
        out += texts.FINANCE_SUMMARY_EXPENSE.format(total=summary["expense_total"])
        for row in summary["by_category"]["expense"]:
            cat_info = texts.FINANCE_CATEGORIES.get(row["category"], ("📋", row["category"]))
            out += texts.FINANCE_SUMMARY_CAT_LINE.format(
                emoji=cat_info[0], category=cat_info[1],
                total=row["total"], cnt=row["cnt"]
            )

    net = summary["net"]
    if net > 0:
        out += texts.FINANCE_SUMMARY_NET_POS.format(net=net)
    elif net < 0:
        out += texts.FINANCE_SUMMARY_NET_NEG.format(net=net)
    else:
        out += texts.FINANCE_SUMMARY_NET_ZERO

    out += texts.FINANCE_SUMMARY_BALANCE.format(balance=get_finance_balance(me["id"]))
    await message.answer(out)


# ===== Excel hisobot =====

@router.message(F.text == texts.BTN_FINANCE_EXCEL)
async def finance_excel(message: Message):
    me = get_employee_by_telegram_id(message.from_user.id)
    if not _can_use_finance(me):
        await message.answer(texts.FINANCE_NO_PERMISSION)
        return

    now = tz_now()
    entries = get_monthly_finance_entries(me["id"], now.year, now.month)
    if not entries:
        await message.answer(texts.FINANCE_EXCEL_EMPTY)
        return

    summary = get_monthly_finance_summary(me["id"], now.year, now.month)
    opening = get_finance_balance_before(me["id"], now.year, now.month)
    file_bytes = await asyncio.to_thread(
        _build_finance_excel, entries, summary, now.year, now.month,
        me["full_name"], opening
    )
    filename = f"moliya_{now.year}_{now.month:02d}_{me['full_name'].split()[0]}.xlsx"
    await message.answer_document(
        BufferedInputFile(file_bytes, filename=filename),
        caption=f"📥 Moliya hisoboti — {texts.MONTHS_UZ[now.month]} {now.year}"
    )


def _build_finance_excel(entries, summary, year: int, month: int,
                         owner_name: str, opening: int = 0) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Yozuvlar"

    title = ws.cell(row=1, column=1,
                    value=f"💰 Moliya — {owner_name} · {texts.MONTHS_UZ[month]} {year}")
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.row_dimensions[1].height = 22

    headers = ["Sana", "Vaqt", "Kategoriya", "Izoh",
               "Rasxod (so'm)", "Kirim (so'm)", "Qoldiq (so'm)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Side(border_style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    # Boshlang'ich qoldiq qatori (oldingi oydan)
    pm = month - 1 if month > 1 else 12
    open_row = 4
    open_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    ws.cell(row=open_row, column=3,
            value=f"📦 {texts.MONTHS_UZ[pm]} oyidan qolgan qoldiq")
    oc = ws.cell(row=open_row, column=7, value=opening)
    oc.number_format = '#,##0'
    oc.font = Font(bold=True)
    for c in range(1, 8):
        cell = ws.cell(row=open_row, column=c)
        cell.border = border
        cell.fill = open_fill

    # Yozuvlar: har qator o'z qoldig'i bilan (Python da hisoblanadi, formula emas)
    start = open_row + 1
    running = opening
    exp_total = 0
    inc_total = 0
    for idx, e in enumerate(entries):
        i = start + idx
        amount = e["amount"]
        cat_info = texts.FINANCE_CATEGORIES.get(e["category"], ("📋", e["category"]))
        ws.cell(row=i, column=1, value=fmt_local(e["entry_date"], "%d.%m.%Y"))
        ws.cell(row=i, column=2, value=fmt_local(e["entry_date"], "%H:%M"))
        ws.cell(row=i, column=3, value=f"{cat_info[0]} {cat_info[1]}")
        # Izoh (avans bo'lsa xodim nomi bilan)
        note = e["note"] or ""
        if dict(e).get("linked_employee_id"):
            emp = get_employee_by_id(e["linked_employee_id"])
            if emp:
                note = f"Avans: {emp['full_name']}" + (f" — {note}" if note else "")
        ws.cell(row=i, column=4, value=note)
        if e["entry_type"] == "expense":
            ac = ws.cell(row=i, column=5, value=amount)
            ac.number_format = '#,##0'
            ac.font = Font(color="8B0000")
            running -= amount
            exp_total += amount
        else:
            ac = ws.cell(row=i, column=6, value=amount)
            ac.number_format = '#,##0'
            ac.font = Font(color="006400")
            running += amount
            inc_total += amount
        qc = ws.cell(row=i, column=7, value=running)
        qc.number_format = '#,##0'
        qc.font = Font(bold=True)
        for c in range(1, 8):
            ws.cell(row=i, column=c).border = border

    # JAMI qatori
    last = start + len(entries) - 1
    trow = last + 1
    ws.cell(row=trow, column=4, value="JAMI").font = Font(bold=True)
    te = ws.cell(row=trow, column=5, value=exp_total)
    te.number_format = '#,##0'
    te.font = Font(bold=True, color="8B0000")
    ti = ws.cell(row=trow, column=6, value=inc_total)
    ti.number_format = '#,##0'
    ti.font = Font(bold=True, color="006400")
    tq = ws.cell(row=trow, column=7, value=running)
    tq.number_format = '#,##0'
    tq.font = Font(bold=True)
    for c in range(1, 8):
        ws.cell(row=trow, column=c).border = border
        ws.cell(row=trow, column=c).fill = open_fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 16

    # Sheet 2: Xulosa
    ws2 = wb.create_sheet("Xulosa")
    ws2.cell(row=1, column=1, value="Tur").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Turkum").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Yozuvlar").font = Font(bold=True)
    ws2.cell(row=1, column=4, value="Jami (so'm)").font = Font(bold=True)
    for c in range(1, 5):
        ws2.cell(row=1, column=c).fill = header_fill
        ws2.cell(row=1, column=c).font = Font(bold=True, color="FFFFFF")
        ws2.cell(row=1, column=c).border = border

    row = 2
    for tp in ("income", "expense"):
        type_label = "Kirim" if tp == "income" else "Chiqim"
        for r in summary["by_category"].get(tp, []):
            cat_info = texts.FINANCE_CATEGORIES.get(r["category"], ("📋", r["category"]))
            ws2.cell(row=row, column=1, value=type_label)
            ws2.cell(row=row, column=2, value=f"{cat_info[0]} {cat_info[1]}")
            ws2.cell(row=row, column=3, value=r["cnt"])
            ac = ws2.cell(row=row, column=4, value=r["total"])
            ac.number_format = '#,##0'
            ac.font = Font(color="006400" if tp == "income" else "8B0000")
            for c in range(1, 5):
                ws2.cell(row=row, column=c).border = border
            row += 1

    row += 1
    ws2.cell(row=row, column=1, value="Jami kirim").font = Font(bold=True)
    c = ws2.cell(row=row, column=4, value=summary["income_total"])
    c.number_format = '#,##0'; c.font = Font(bold=True, color="006400")
    row += 1
    ws2.cell(row=row, column=1, value="Jami chiqim").font = Font(bold=True)
    c = ws2.cell(row=row, column=4, value=summary["expense_total"])
    c.number_format = '#,##0'; c.font = Font(bold=True, color="8B0000")
    row += 1
    ws2.cell(row=row, column=1, value="Sof natija").font = Font(bold=True, size=12)
    c = ws2.cell(row=row, column=4, value=summary["net"])
    c.number_format = '#,##0'
    c.font = Font(bold=True, color="006400" if summary["net"] >= 0 else "8B0000")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
